# -*- coding: utf-8 -*-
"""
Created on Tue Aug 10 15:3:44 2021

Se calcula el pronóstico de cada para _-sucursal

@author: sebastian
"""
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
import datetime
import matplotlib.pyplot as plt
from dateutil.relativedelta import relativedelta

from sklearn.svm import SVR
from sklearn.svm import LinearSVR
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler, MinMaxScaler, MaxAbsScaler, RobustScaler

from sklearn.ensemble import RandomForestRegressor
from sklearn.datasets import make_regression
from sklearn.linear_model import LinearRegression

from openpyxl import load_workbook

from scipy.stats import pearsonr
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score
from sklearn.metrics import mean_squared_error
from sklearn.preprocessing import LabelEncoder
import statsmodels.api as sm
import statsmodels.formula.api as smf

from statsmodels.tsa.api import SimpleExpSmoothing,Holt,ExponentialSmoothing
from statsmodels.graphics.tsaplots import plot_acf


from kats.consts import TimeSeriesData
from kats.models.prophet import ProphetModel, ProphetParams
from kats.models.sarima import SARIMAModel, SARIMAParams


from sklearn.model_selection import cross_val_score
from sklearn.model_selection import RepeatedStratifiedKFold

from lightgbm import LGBMRegressor
import os
import ast

import time

import seaborn as sns
import matplotlib.dates as mdates

from hampel import hampel

import sys
sys.path.insert(0, '/Documentos/PROJECT/Cooprinsem2022/Cooprinsem2/Prototipo_v6')

#from ConnectionSQL import*

class Forescast():
    def __init__(self, 
                 levelconfidence_a = .98,
                 levelconfidence_b = .95,
                 levelconfidence_c = .90,
                 acceptance_tol_ub = .3,
                 acceptance_tol_lb = .3,
                 reviewperiod      = 7,
                 name=None):
        
         self.par = name
         # Tolerancia de aceptación de pronóstico
         self.acceptance_tol_ub    = acceptance_tol_ub
         self.acceptance_tol_lb    = acceptance_tol_lb
         self.reviewperiod = reviewperiod
         self.t0 = time.time()
         
         # Nivel de confianza por ABC
         self.levelconfidence_a = levelconfidence_a
         self.levelconfidence_b = levelconfidence_b
         self.levelconfidence_c = levelconfidence_c
         
         # Directorio de Base de Datos Data Interna
         self.datafile ='Data/Predict/DataPredict_%s.csv'%self.par
         
         
         # Directorio de Base de Datos Data Exogena
         self.path_varexog = 'Data/PredictVarsExogenasv4.csv'
         
         # Directorio de algoritmos
         self.path_rf = 'Algorithms/Parameters/RandomForest_complete.csv'
         self.path_lbgm ='Algorithms/Parameters/lgbmtest_complete.csv'
         self.path_svm = 'Algorithms/Parameters/svmtest_complete.csv'
         self.path_sarima = 'Algorithms/Parameters/sarima_complete.csv'
         #self.path_HoltWinter = 'Algorithms/Parameters/HoltWinter_complete.csv'
         
         # Columnas que se usarán como predidctores
         self.weather       = ['precip_0trim_mean','precip_1trim_mean',
                               'dia_lluvia_0trim_mean','dia_lluvia_1trim_mean',
                               ]
         self.temperature  =[ 'tmean_0trim_mean','tmean_1trim_mean',
                             ]
         
         self.hotday = ['dia_caluroso_0trim_mean','dia_caluroso_1trim_mean']
         
         self.coldday = ['dia_helada_0trim_mean','dia_helada_1trim_mean']
         
         self.macroeconomic = ['Dolar','IPC','Desempleo','IMACEC']
         
         self.priceproduct = ['Precio_carne_0trim_mean','Precio_carne_1trim_mean',
                               'Precio_leche_0trim_mean','Precio_leche_1trim_mean'
                               ]
                  
         self.dictpredictors = {
                                 'all'    : self.weather+ self.macroeconomic + self.hotday + self.coldday + self.priceproduct,
                                'weather-all' : self.weather + self.hotday + self.coldday,
                                'macroeconomic-all' : self.macroeconomic + self.priceproduct,
                                'timedate' : []
                          }

         self.dfpredictors = pd.read_csv(self.path_varexog, sep=',',
                                         decimal='.', index_col=0)
                  
    def read(self, filter):
        t0 = time.time()
        
        with open('ReportScript/Forecast/Estado_ejecución_Pronóstico_%s.txt'%self.par, 'w') as f:
            f.write('Inciando carga de datos \n')
        f.close()
        
        #%% Dataframe GPI Ventas--------------------------------
        self.df             = pd.read_csv(self.datafile, sep=",", decimal='.', index_col=0 )
        #self.df.to_csv('Results/Agregado/read_%s.csv'%self.par)
        self.dfumb = self.df[['IdMaterial','UMBase']]
        self.dfumb = self.dfumb.drop_duplicates()
        self.dfumb = self.dfumb.set_index('IdMaterial')
                
        filterdata = filter
        
        self.df['Linea'] = self.df['IdMaterial'].map(str)
        self.df['Linea'] = self.df['Linea'].str.slice(0, 2)
                                                          
        self.ModelParamsRF    = pd.read_csv(self.path_rf, sep=",", decimal='.', index_col=0 )
        self.ModelParamsLGBM  = pd.read_csv(self.path_lbgm, sep=",", decimal='.', index_col=0 )
        self.ModelParamsSVM   = pd.read_csv(self.path_svm , sep=",", decimal='.', index_col=0 )
        self.ModelParamsSarima   = pd.read_csv(self.path_sarima, sep=",", decimal='.', index_col=0 )
        
        self.df['IdPeriodo'] = pd.to_datetime(self.df['IdPeriodo'], format="%Y/%m")
        
        lastdate            = self.df['IdPeriodo'].max()
        
        self.year, self.month = int(lastdate.year), int(lastdate.month) #2021,3
        
        self.date_start     = datetime.date(self.year, self.month, 3)
        
                # Considera el útlimo mes completado        
        current_date = datetime.datetime.today()
        #if ((current_date.year == self.date_start.year) and
         #  (current_date.month == self.date_start.month)):
        self.date_start = self.date_start - relativedelta(months=1)
        self.df = self.df[self.df['IdPeriodo'] < str(self.date_start)]
        # self.df.to_csv('Results/Agregado/read_%s.csv'%self.par)       
        self.date_start_predict = datetime.date(self.year, self.month, 1) 
        self.period         = 6
        self.date_finish    = self.date_start + relativedelta(months=self.period)
        
        self.date_lastyear  = self.date_start - relativedelta(months=12)
        
        getdate = lambda x : self.date_start + relativedelta(months=x)
        getdate2 = lambda x : self.date_lastyear + relativedelta(months=x)
        
        self.period_proy  = ['DdaProy %s/%s'%(getdate(i).year,getdate(i).month) for i in range(1,self.period+1)]
        
        print(self.period_proy)
        
        self.lastyear_name = ['%s/%s'%(getdate2(i).year,getdate2(i).month) for i in range(1,13)]
        

        self.trim_name = ['Trim %s/%s - %s/%s'%(getdate2(i).year,getdate2(i).month,getdate2(i+2).year,getdate2(i+2).month) 
                          for i in range(1,13,3)]
        
        self.trim_proy  = ['Trim Proy %s/%s - %s/%s'%(getdate(i).year,getdate(i).month,getdate(i+2).year,getdate(i+2).month) 
                           for i in range(1,self.period+1,3)]
                


        self.predict_date = [datetime.date(getdate(i).year, getdate(i).month, 1) for i in range(1,self.period+1)]
        
        self.date_start, self.date_finish, self.date_lastyear = str(self.date_start), str(self.date_finish), str(self.date_lastyear)

        dfaux                 = self.df[(self.df['IdPeriodo'] <= self.date_start) & 
                                        (self.df['IdPeriodo'] > self.date_lastyear)]
        

        def PeriodColumn(x):
            return '%s/%s'%(x.year,x.month)
        
        dfaux['IdPeriodo'] = pd.DataFrame(dfaux.apply(lambda x: PeriodColumn(x['IdPeriodo']),
                                                                  axis=1).tolist(), index=dfaux.index)
    
        self.dflastyear       = pd.pivot_table(dfaux,
                                                values  = 'CtdadUMBase',
                                                columns = 'IdPeriodo',
                                                index   = ['IdMaterial', 'IdCeSum'],
                                                fill_value = 0).reset_index()
        

        
        dfaux['CtdadUMBase'] = pd.to_numeric(dfaux['CtdadUMBase'])
        dfaux                 = dfaux.groupby(['IdMaterial',
                                                'IdCeSum'
                                                ]).agg({
                                                'CtdadUMBase':'sum'}).reset_index()
        dfaux2                 = dfaux.groupby(['IdMaterial',
                                                ]).agg({
                                                'CtdadUMBase':'sum'}).reset_index()
        dfaux['MatSuc']       = dfaux['IdMaterial'].map(str) + '-' + dfaux['IdCeSum']
        Active                = dfaux[dfaux['CtdadUMBase'] > 0.0]['IdMaterial'].unique()
        
        self.df['MatSuc']     = self.df['IdMaterial'].map(str) + '-' + self.df['IdCeSum']
        self.df               = self.df[self.df['IdMaterial'].isin(Active)]
        # self.df.to_csv('Results/Agregado/read_%s.csv'%self.par)
        self.df.drop(['MatSuc'], inplace=True, axis=1)
        # self.df.to_csv('Results/Agregado/read_%s.csv'%self.par)
        dfsucprct = dfaux.groupby(['IdMaterial','IdCeSum']).agg({'CtdadUMBase':'sum'}).reset_index()
        
        dfsucprct['Porcentaje'] = dfsucprct['CtdadUMBase']/ dfaux.groupby('IdMaterial')['CtdadUMBase'].transform('sum')
        
        self.dflastyear = self.dflastyear.merge(dfsucprct[['IdMaterial','IdCeSum','Porcentaje']],
                                                on = ['IdMaterial','IdCeSum'], how='left',
                                                )
                                
        dfM                 = pd.read_excel('Data/Lead_Time_Linea14.xlsx', index_col=0)
        dfM                 = dfM[['IdMaterial','IdCeSum','Provision','LeadTime']] 
        dfM['IdMaterial']   = dfM['IdMaterial'].map(int)
        dfM['LeadTime']     = dfM['LeadTime'].round(0)
        dfM['Provision']     = dfM['Provision'].str.strip()
        dfM['IdCeSum']     = dfM['IdCeSum'].str.strip()       
        
        self.dflastyear = self.dflastyear.merge(dfM,
                                                how='left',
                                                on=['IdMaterial','IdCeSum'])
        
        self.dflastyear['Provision'] = self.dflastyear['Provision'].fillna("Directo")
        
        self.dfabc               = pd.read_csv('Data/ABCxSucursal.csv', sep=",", decimal='.' )
        self.dfabc               = self.dfabc[['IdMaterial','Sucursal','IdCeSum','ABC_Sucursal','ABC','Margen_Total','Margen_Suc']]

        self.dfabc['IdCeSum']     = self.dfabc['IdCeSum'].str.strip()
        
        self.dflastyear = self.dflastyear.merge(self.dfabc, how='left', on = ['IdMaterial' , 'IdCeSum'])  
        
        self.dflastyear['ABC_Sucursal'] = self.dflastyear['ABC_Sucursal'].fillna('B')
        self.dflastyear['ABC'] = self.dflastyear['ABC'].fillna('B')
        
        #self.dflastyear['LeadTime'] = self.dflastyear['LeadTime'].fillna(90)
                
        print(self.date_start)
        print(self.date_finish)
        print(self.df.IdPeriodo.max())
               
        with open('ReportScript/Forecast/Estado_ejecución_Pronóstico_%s.txt'%self.par, 'a') as f:
            f.write('Fecha de inicio de pronóstico: %s \n'%self.date_start)
            f.write('Fecha de termino de pronóstico: %s \n'%self.date_finish)
            f.write('Ha finalizado la carga de datos: %s(s), duración actual: %s(s) \n'%(time.time() - t0,time.time()-self.t0))
        f.close()

    def updateFormat(self):
        # Transformar datos a númerico
        # self.df.to_csv('Results/Agregado/read_%s.csv'%self.par)
        self.df["IdPeriodo"] = pd.to_datetime(self.df['IdPeriodo'], format="%Y/%m")
                
        self.df['CtdadUMBase'] = pd.to_numeric(self.df['CtdadUMBase'])
        
        self.df.rename(columns={"CtdadUMBase": "Demanda"}, inplace=True)
        
        self.df['Demanda'] = self.df['Demanda'].map(float)
        
        self.df = self.df.fillna(0)
                
        IdCeSumes = list(self.df['IdCeSum'].unique())
        IdMaterial = list(self.df['IdMaterial'].unique())
                
        self.df['Demanda']          = self.df['Demanda'].map(float)
        self.df['MesPeriodo']       = self.df['MesPeriodo'].map(int) 
        self.df['Dolar']            = self.df['Dolar'].map(float)
        self.df['IMACEC']           = self.df['IMACEC'].map(float)
        self.df['Desempleo']        = self.df['Desempleo'].map(float)
        self.df['IPC']              = self.df['IPC'].map(float)
        
        self.df['precip_1trim_mean'] = self.df['precip_1trim_mean'].map(float)
        self.df['precip_0trim_mean'] = self.df['precip_0trim_mean'].map(float)

        self.df['dia_lluvia_1trim_mean'] = self.df['dia_lluvia_1trim_mean'].map(float)
        self.df['dia_lluvia_0trim_mean'] = self.df['dia_lluvia_0trim_mean'].map(float)

        self.df['dia_caluroso_1trim_mean'] = self.df['dia_caluroso_1trim_mean'].map(float)
        self.df['dia_caluroso_0trim_mean'] = self.df['dia_caluroso_0trim_mean'].map(float)

        self.df['dia_helada_1trim_mean'] = self.df['dia_helada_1trim_mean'].map(float)
        self.df['dia_helada_0trim_mean'] = self.df['dia_helada_0trim_mean'].map(float)

        self.df['tmean_1trim_mean'] = self.df['tmean_1trim_mean'].map(float)
        self.df['tmean_0trim_mean'] = self.df['tmean_0trim_mean'].map(float)
        
        self.df['Precio_carne_1trim_mean'] = self.df['Precio_carne_1trim_mean'].map(float)
        self.df['Precio_carne_0trim_mean'] = self.df['Precio_carne_0trim_mean'].map(float)
        
        self.df['Precio_leche_1trim_mean'] = self.df['Precio_leche_1trim_mean'].map(float)
        self.df['Precio_leche_0trim_mean'] = self.df['Precio_leche_0trim_mean'].map(float)
        
        self.IdMaterialFilter = list(self.df['IdMaterial'].unique())
                
        self.df = self.df.sort_values(by='IdPeriodo')
        
        self.IdMaterial = self.df['IdMaterial'].unique()

        self.df['IdMaterial']            = self.df['IdMaterial'].map(int) 
        self.ModelParamsSVM['SKU']   = self.ModelParamsSVM['SKU'].map(int) 
        self.ModelParamsRF['SKU']   = self.ModelParamsRF['SKU'].map(int) 
        self.ModelParamsLGBM['SKU']   = self.ModelParamsLGBM['SKU'].map(int) 
        self.ModelParamsSarima['SKU']   = self.ModelParamsSarima['SKU'].map(int) 
        
        self.ModelParamsRF.set_index(['SKU'], inplace = True)
        self.ModelParamsLGBM.set_index(['SKU'], inplace = True)
        self.ModelParamsSarima.set_index(['SKU'], inplace = True)
        self.ModelParamsSVM.set_index(['SKU'], inplace = True)
                        
        self.dfpredictors["IdPeriodo"] = pd.to_datetime(self.dfpredictors['IdPeriodo'], format="%Y/%m")
        self.dfpredictors['MesPeriodo'] = self.dfpredictors['MesPeriodo'].map(int) 
      
    #%%  
    def Randomforest(self, train, test, params):

        RMSE = np.inf
        
        preproc = [StandardScaler(),
                   MaxAbsScaler(), 
                   MinMaxScaler(),
                   RobustScaler(quantile_range = (0.1,0.9))
                  ]
        
        test.MesPeriodo = test.MesPeriodo.map(str)
        train.MesPeriodo = train.MesPeriodo.map(str)
        
        regr = RandomForestRegressor(max_samples       =(None
                                                         if params[0] == 'None' 
                                                         else float(params[0])),
                                                         max_features      =  (None
                                                                              if params[1] == 'None' 
                                                                              else params[1]),
                                                         n_estimators      = int(params[2]),
                                                         max_depth         = (None
                                                                              if params[3] == 'None' 
                                                                              else float(params[3])),
                                                         bootstrap         = True if params[4] == 'True' 
                                                                             else False,
                                                         n_jobs            = 1,
                                                         random_state      = 123)
        scaler = preproc[int(params[7])]
        for keys_ in self.dictpredictors.keys():    
            predictors_ = self.dictpredictors[keys_]
            
            predictors2_ = predictors_ #+ branches 
                                                                 
            X, y = train[predictors2_], train['Demanda']  
            
            X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=0, test_size=0.2)
            
            X_train = scaler.fit_transform(X_train)
            X_test = scaler.transform(X_test)
            
            regr.fit(X_train, y_train)
                                                
            y_aux1 = regr.predict(X_test)
            
            y_pred = [max(0,y_aux1[i]) for i in range(len(y_aux1))]
                                
            
            RMSE_ = np.mean(np.abs(y_pred-y_test)/(1e-3+np.abs(y_test)+np.abs(y_pred)))
            
            if RMSE_ < RMSE:
                RMSE = RMSE_
                predictors = predictors2_
                y_pred_best = y_pred

        RMSE, MAE, MAPE = self.metric(y_test, y_pred_best)

        X2 = test[predictors]
        X  =  train[predictors]
        
        X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=0, test_size=0.2)

        X3 = X.append(X2)
        
        X_train = scaler.fit_transform(X_train)
        X2 = scaler.transform(X2)
        X3 = scaler.transform(X3)        
        
        regr.fit(X_train, y_train)
        
        y_aux1 = regr.predict(X2)
        y_aux2 = regr.predict(X3)
                
        y_pred = [max(0,y_aux1[i]) for i in range(len(y_aux1))]
        y_pred2 = [max(0,y_aux2[i]) for i in range(len(y_aux2))]
                                                            
        return y_pred , RMSE, MAE, MAPE, y_pred2
    
    #%%  
    def SVM(self, train, test, params):

        RMSE = np.inf
        preproc = [StandardScaler(),
                   MaxAbsScaler(), 
                   MinMaxScaler(),
                   RobustScaler(quantile_range = (0.1,0.9))
                  ]
        
        test.MesPeriodo = test.MesPeriodo.map(str)
        train.MesPeriodo = train.MesPeriodo.map(str)

        regr = make_pipeline(preproc[int(params[0])],
                             SVR(
                                 C         = float(params[1]),
                                 epsilon   = (None
                                             if params[2] == 'None' 
                                             else float(params[1])),
                                 gamma     = params[3],
                                 shrinking = (params[4] == 'True'),
                                 kernel    = params[5],
                                 verbose = True,
                                 max_iter=300,
                                 ),
                                )

        for keys_ in self.dictpredictors.keys():    
            predictors2_ = self.dictpredictors[keys_]
                                                                             
            X, y = train[predictors2_], train['Demanda']  
            
            X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=0, test_size=0.2)
           
            regr.fit(X_train, y_train)
                                                
            y_aux1 = regr.predict(X_test)
            
            y_pred = [max(0,y_aux1[i]) for i in range(len(y_aux1))]
            

            RMSE_ = np.mean(np.abs(y_pred-y_test)/(1e-3+np.abs(y_test)+np.abs(y_pred)))
            
            if RMSE_ < RMSE:
                RMSE = RMSE_
                predictors = predictors2_
                y_pred_best = y_pred
                

        RMSE, MAE, MAPE = self.metric(y_test, y_pred_best)

        X2 = test[predictors]
        X  =  train[predictors]
        
        X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=0, test_size=0.2)

        X3 = X.append(X2)
                
        regr.fit(X_train, y_train)
        
        y_aux1 = regr.predict(X2)
        y_aux2 = regr.predict(X3)
        
        y_pred = [max(0,y_aux1[i]) for i in range(len(y_aux1))]
        y_pred2 = [max(0,y_aux2[i]) for i in range(len(y_aux2))]
                            
        return y_pred, RMSE, MAE, MAPE, y_pred2

    def lgbm(self, train, test, params):
        
        RMSE = np.inf
        
        preproc = [StandardScaler(),
                   MaxAbsScaler(), 
                   MinMaxScaler(),
                   RobustScaler(quantile_range = (0.1,0.9))
                  ]

        regr = LGBMRegressor(
            n_estimators            = 1000,#int(params[0]), # equivalent num_iterations
            max_depth               = int(params[0]),#3,,     #(importante hasta 8 (más chico mejora))
            learning_rate           = float(params[1]),#0.1515,#float(params[2]),(importante)
            n_jobs                  = 1,
            min_gain_to_split       = float(params[2]),#0.25,
            min_data_in_leaf        = 2,
            lambda_l1               = float(params[3]),#100, #(importante (más alto mejora))
            lambda_l2               = float(params[4]),#0, #(imporntante)
            seed = 123
            )

        scaler = preproc[int(params[5])]
        for keys_ in self.dictpredictors.keys():    
            predictors_ = self.dictpredictors[keys_]
            
            predictors2_ = predictors_ #+ branches 
                                                                 
            X, y = train[predictors2_], train['Demanda']  
            
            X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=0, test_size=0.2)
            
            X_train = scaler.fit_transform(X_train)
            X_test = scaler.transform(X_test)
            
            regr.fit(X_train, y_train)
                                                
            y_aux1 = regr.predict(X_test)
            
            y_pred = [max(0,y_aux1[i]) for i in range(len(y_aux1))]
                                
            
            RMSE_ = np.mean(np.abs(y_pred-y_test)/(1e-3+np.abs(y_test)+np.abs(y_pred)))
            
            if RMSE_ < RMSE:
                RMSE = RMSE_
                predictors = predictors2_
                y_pred_best = y_pred

        RMSE, MAE, MAPE = self.metric(y_test, y_pred_best)

        X2 = test[predictors]
        X  =  train[predictors]
        
        X_train, X_test, y_train, y_test = train_test_split(X, y, random_state=0, test_size=0.2)

        X3 = X.append(X2)
        
        X_train = scaler.fit_transform(X_train)
        X2 = scaler.transform(X2)
        X3 = scaler.transform(X3)        
        
        regr.fit(X_train, y_train)
        
        y_aux1 = regr.predict(X2)
        y_aux2 = regr.predict(X3)
                
        y_pred = [max(0,y_aux1[i]) for i in range(len(y_aux1))]
        y_pred2 = [max(0,y_aux2[i]) for i in range(len(y_aux2))]
                
        return y_pred , RMSE, MAE, MAPE, y_pred2

        
    def SARIMA(self, train, test, params):
        #  Construct TimeSeriesData object        
        train = train.rename(columns={"IdPeriodo": "time", "Demanda": "value"})
        train = train[['time','value']]

        train = train.sort_values(by='time')
        train_ = train[:int(len(train)*.8)] 
        test_ = train[int(len(train)*.8):] 
                
        ts = TimeSeriesData(train_)
        # create SARIMA param class
        params_sarima =  SARIMAParams(p              = int(params[0]),#int(params['p']),
                                      d              = int(params[1]),#int(params['d']),
                                      q              = int(params[2]),#int(params['q']),
                                      # exog=month,
                                      seasonal_order = (int(params[3]),int(params[4]),
                                                        int(params[5]),int(params[6])),
                                      trend           = params[7],
                                      enforce_stationary = False,
                                      enforece_invertibility=False,
                                      # freq = 'M'
                                      
                                              )            
        # initiate SARIMA model
        m = SARIMAModel(data=ts,  params=params_sarima)
        m.fit(maxiter=100, full_output = False)
        sarima_ = m.predict(steps = len(test_), params = params_sarima)
        y_pred = np.array([max(value,0) for value in sarima_['fcst']])
        y_test = np.array([value for value in test_['value']])
        
        RMSE, MAE, MAPE = self.metric(y_test, y_pred)        
        
        train = train#[(train['time'] < self.date_start)]
        
        ts = TimeSeriesData(train)
        m = SARIMAModel(data=ts, params=params_sarima)
        
        m.fit()
        sarima = m.predict(steps = self.period, freq="MS")
        
        y_aux1 = sarima['fcst']
        y_aux1 = [value for value in y_aux1]
        
        y_aux2 = sarima['fcst_upper']
        y_aux2 = [value for value in y_aux2]
        
        y_pred = []
        for i in range(len(y_aux1)):
            if y_aux1[i] < 0:
                y_pred.append(0)#max(0,y_aux2[i]))
            else:
                y_pred.append(y_aux1[i])
            
        y_pred = [value for value in y_pred]
        
        # Get history + forecast
        sarima = m.predict(steps = self.period, freq="MS", include_history = True)
        
        y_aux1 = sarima['fcst']
        y_aux1 = [value for value in y_aux1]
        
        y_aux2 = sarima['fcst_upper']
        y_aux2 = [value for value in y_aux2]
        
        y_pred2 = []
        for i in range(len(y_aux1)):
            if y_aux1[i] < 0:
                y_pred2.append(0)#max(0,y_aux2[i]))
            else:
                y_pred2.append(y_aux1[i])
            
        y_pred2 = [value for value in y_pred2]
        
        return y_pred, RMSE, MAE, MAPE, y_pred2
    
    def HoltWinteradd(self, train_, test_, params):
        
        #train_ es la variable que contiene la data completa
        
        
        total_rows = len(train_.axes[0])  #obtiene total de columnas
        train = train_[0:int(total_rows*0.8)]
        train.index = train.IdPeriodo
        test = train_[int(total_rows*0.8):]
        test.index = test.IdPeriodo


        #plotting model result
        fit1 = ExponentialSmoothing(np.asarray(train['Demanda']),seasonal_periods=12,trend='add',seasonal='add').fit()
        y_pred = fit1.forecast(len(test))
        y_pred = np.array([max(value,0) for value in y_pred])
        y_test = np.array([value for value in test['Demanda']])
        
        RMSE, MAE, MAPE = self.metric(y_test, y_pred)
        
        train_.index = train_.IdPeriodo
        fit2 = ExponentialSmoothing(np.asarray(train_['Demanda']),seasonal_periods=12,trend='add',seasonal='add').fit()
        y_pred2=fit2.forecast(self.period)
        y_pred=fit2.fittedvalues
        y_pred2 = np.array([max(value,0) for value in y_pred2])
        y_pred=np.append(y_pred,[max(value,0) for value in y_pred2])
        # print(len(y_pred),len(y_pred2))
        
        
        return y_pred2, RMSE, MAE, MAPE, y_pred
    
    def HoltWintermul(self, train_, test_, params):
        #train_ es la variable que contiene la data completa
        
        
        total_rows = len(train_.axes[0])  #obtiene total de columnas
        train = train_[0:int(total_rows*0.8)]
        train.index = train.IdPeriodo
        test = train_[int(total_rows*0.8):]
        test.index = test.IdPeriodo


        #plotting model result
        fit1 = ExponentialSmoothing(np.asarray(train['Demanda']),seasonal_periods=12,trend='mul',seasonal='mul').fit()
        y_pred = fit1.forecast(len(test))
        y_pred = np.array([max(value,0) for value in y_pred])
        y_test = np.array([value for value in test['Demanda']])
        
        RMSE, MAE, MAPE = self.metric(y_test, y_pred)
        
        train_.index = train_.IdPeriodo
        fit2 = ExponentialSmoothing(np.asarray(train_['Demanda']),seasonal_periods=12,trend='mul',seasonal='mul').fit()
        y_pred2=fit2.forecast(self.period)
        y_pred=fit2.fittedvalues
        y_pred2 = np.array([max(value,0) for value in y_pred2])
        y_pred=np.append(y_pred,[max(value,0) for value in y_pred2])
        # print(len(y_pred),len(y_pred2))
        
        
        return y_pred2, RMSE, MAE, MAPE, y_pred

            
    
    def SaveSheet(self, dfsave, namesheet):
                                 
        path = r"Results/DataPredict%s.xlsx"%self.par
        try:
            book = load_workbook(path)
            writer = pd.ExcelWriter(path, engine = 'openpyxl')
            writer.book = book
            sheets = book.sheetnames
            if namesheet in sheets:
                std=book[namesheet]
                book.remove(std)
                dfsave.to_excel(writer, sheet_name=namesheet)
                writer.save()
                writer.close()
            else:
                dfsave.to_excel(writer, sheet_name=namesheet)
                writer.save()
                writer.close()
        except:
            with pd.ExcelWriter(path) as writer:
                dfsave.to_excel(writer, sheet_name=namesheet)
                
    def algorithm(self, NameAlg,params,train,test):
        
        if NameAlg == 'rforest':
           predict, RMSE, MAE, MAPE, hist_predict = self.Randomforest(train, test,  params)
           predict = list(predict)
           ddamean = [np.mean(np.array(predict))]

        elif NameAlg == 'svm' :
           predict, RMSE, MAE, MAPE, hist_predict = self.SVM(train, test, params)
           predict = list(predict)
           ddamean = [np.mean(np.array(predict))]
           
        elif NameAlg == 'lgbm' :
           predict, RMSE, MAE, MAPE, hist_predict = self.lgbm(train, test, params)
           predict = list(predict)
           ddamean = [np.mean(np.array(predict))]
         
        elif NameAlg == 'sarima' :
            predict, RMSE, MAE, MAPE, hist_predict = self.SARIMA(train, test, params)
            predict = list(predict)
            ddamean = [np.mean(np.array(predict))]
            
        elif NameAlg == 'HoltWinteradd' :
            predict, RMSE, MAE, MAPE, hist_predict = self.HoltWinteradd(train, test, params)
            predict = list(predict)
            ddamean = [np.mean(np.array(predict))]
        
        elif NameAlg == 'HoltWintermul' :
            predict, RMSE, MAE, MAPE, hist_predict = self.HoltWintermul(train, test, params)
            predict = list(predict)
            ddamean = [np.mean(np.array(predict))] 
            
        else:
            print('FALLA')
            
        return predict, RMSE, MAE, MAPE, ddamean, hist_predict
    
    def metric(self, y_test, y_pred):
        SMAPE = np.mean(np.abs(y_pred-y_test)/(1e-3+np.abs(y_test)+np.abs(y_pred)))
        # Mean Absolute Error
        MAE = np.mean(np.abs(y_pred-y_test))
        # Root mean squared error 
        RMSE = np.sqrt(np.mean((y_pred-y_test)**2))
        
        return RMSE, MAE, SMAPE
    
    def seleccionarPeriodos(self, df):
        
        minPeriodo = min(df['Periodo'])
        maxPeriodo = max(df['Periodo'])
        cantPeriodos = (maxPeriodo - minPeriodo) + 1
    
        if(cantPeriodos % 2 == 0):
            grupos = cantPeriodos/2
            flag = 0  # periodos exactos
        else:
            grupos = cantPeriodos//2  #--> se guarda la parte entera de la division
            flag = 1  # el ultimo queda solo
     
        return minPeriodo, maxPeriodo, grupos, flag



    def seleccionarIndices(self, df, indices, minPeriodo, maxPeriodo, grupos, flag, contador):
        
        periodo1 = minPeriodo
        periodo2 = periodo1 + 1
        lista = []

        # se verifica si queda el ultimo año suelto o los periodos son exactos
        if(flag != 0):

            if(contador <= grupos):

                if(flag == 1 and periodo1 == maxPeriodo):
                    # extrar solo hasta [periodo1]
                    for i in range(len(df)):
                        if(df.iloc[i]['Periodo'] == periodo1):
                            lista.append(df.iloc[i]['position'])

                    tramoMin = min(lista)
                    tramoMax = max(lista)+1
                    
                    indices.loc[contador,'tramo'] = contador
                    indices.loc[contador,'indice1'] = tramoMin 
                    indices.loc[contador,'indice2'] = tramoMax

                    self.seleccionarIndices(df, indices, minPeriodo+2, maxPeriodo, grupos, flag, contador+1)

                else:
                    # extraer [periodo1-periodo2]
                    for i in range(len(df)):
                        if(df.iloc[i]['Periodo'] == periodo1 or df.iloc[i]['Periodo'] == periodo2):
                            lista.append(df.iloc[i]['position'])

                    tramoMin = min(lista)
                    tramoMax = max(lista)+1

                    indices.loc[contador,'tramo'] = contador
                    indices.loc[contador,'indice1'] = tramoMin 
                    indices.loc[contador,'indice2'] = tramoMax

                    self.seleccionarIndices(df, indices, minPeriodo+2, maxPeriodo, grupos, flag, contador+1)
        
        else:
            if(contador < grupos):
                for i in range(len(df)):
                    # extraer [periodo1-periodo2]
                    if(df.iloc[i]['Periodo'] == periodo1 or df.iloc[i]['Periodo'] == periodo2):
                            lista.append(df.iloc[i]['position'])
                
                tramoMin = min(lista)
                tramoMax = max(lista)+1

                indices.loc[contador,'tramo'] = contador
                indices.loc[contador,'indice1'] = tramoMin 
                indices.loc[contador,'indice2'] = tramoMax

                self.seleccionarIndices(df, indices, minPeriodo+2, maxPeriodo, grupos, flag, contador+1)

        return indices
    
        
    
    def smooth_out(self, df):
        
        Q1 = df['Demanda'].quantile(0.25)
        Q3 = df['Demanda'].quantile(0.75)
        # print('mostrando Q1:', Q1)
        # print('mostrando Q3:', Q3)
        IQR = Q3-Q1
        # cuando el BI o bigote inferior es negativo o inferior al minimo, el bigote inferior toma el valor del minimo.
        BI_Calculado = (Q1-IQR)
        BS_Calculado = (Q3+IQR)
        # print('Bigote Inferior: ',BI_Calculado)
        # print('Bigote Superior: ',BS_Calculado)
        for i in range(len(df)): 

            if(df.iloc[i]['Demanda'] < BI_Calculado):
                nuevo_BS = BI_Calculado
                df.loc[i, 'outlier'] = nuevo_BS

            elif(df.iloc[i]['Demanda'] > BS_Calculado):
                nuevo_BS = BS_Calculado
                df.loc[i, 'outlier'] = nuevo_BS
            
            else:
                nuevo_BS = df.iloc[i]['Demanda']
                df.loc[i, 'outlier'] = nuevo_BS
                
        return df
    
    
    
    
    def insertarDatos(self, tramo, df):
        
        for i in range(len(tramo)):
            agregar = tramo.iloc[i]['outlier']
            position = int(tramo.iloc[i]['position'])
            df.loc[position,'outlier'] =  agregar
        
        return df
    
    
    
    def initialSmooth(self, df, indices, grupos, flag, i):

        if(flag != 0):
            if(i <= grupos):
                tramo = df[int(indices.iloc[i]['indice1']):int(indices.iloc[i]['indice2'])].reset_index()
                self.smooth_out(tramo)
                self.insertarDatos(tramo, df)
                self.initialSmooth(df, indices, grupos, flag, i+1)

        else:
            if(i < grupos):
                tramo = df[int(indices.iloc[i]['indice1']):int(indices.iloc[i]['indice2'])].reset_index()
                self.smooth_out(tramo)
                self.insertarDatos(tramo, df)
                self.initialSmooth(df, indices, grupos, flag, i+1)

        return df
    
    
    
    
    def general(self, df):
        
        minPeriodo, maxPeriodo, grupos, flag = self.seleccionarPeriodos(df)
        indices  = pd.DataFrame({'tramo': [], 'indice1': [], 'indice2': []})
        indices = self.seleccionarIndices(df, indices, minPeriodo, maxPeriodo, grupos, flag, 0)
        dfFinal = self.initialSmooth(df, indices, grupos, flag, 0)
        
        return dfFinal
    
    def contador(self, df):
        contador = 0
        for i in range(len(df)):
            if df['Demanda'][i] == df ['outlier'][i]:
                contador = contador + 0
            else:
                contador = contador + 1
        
        return contador
                
    def Predict(self):  
        # self.df.to_csv('Results/Agregado/examinando_%s.csv'%self.par)
        t0 = time.time()
        
        with open('ReportScript/Forecast/Estado_ejecución_Pronóstico_%s.txt'%self.par, 'a') as f:
            f.write('Ha iniciado el prónostico \n')
        f.close()
                
        dfsave   = pd.DataFrame()
        default_params = {
                            'sarima':[1,0,0,1,0,1,12,'ct'], 
                            'lgbm'  :[3, 0.29, 0, 0, 1, 0],
                            'svm'   :[0, 1, 0.1, 'scale', 'True', 'rbf'],
                            'rforest':['None', 'auto',100, 'None', 'True', 2, 1, 0],
                            'HoltWinteradd':[0],
                            'HoltWintermul':[0]
                          }
        dfM                 = pd.read_excel('Data/Lead_Time_Linea14.xlsx', index_col=0)
        dfM                 = dfM[['IdMaterial','IdCeSum','Provision','LeadTime']] 
        dfM                 = dfM[dfM['IdMaterial'].isin(self.IdMaterial)]
        dfM['IdMaterial']   = dfM['IdMaterial'].map(int)
        dfM['LeadTime']     = dfM['LeadTime'].round(0)
        dfM['Provision']    = dfM['Provision'].str.strip()
        dfM['IdCeSum']      = dfM['IdCeSum'].str.strip()
        IdMaterialM         = dfM['IdMaterial'].unique()
        
        self.IdMaterial = list(self.IdMaterial)
        
        m_ = 0
                
        lastyearhead = False
                
        self.dfpredictors = self.dfpredictors[self.dfpredictors['CentroSuministrador']=='Osorno']
        self.df = self.df.sort_values(by='IdPeriodo')

        self.df = self.df.groupby(['IdPeriodo',
                                    'MesPeriodo',
                                    'IdMaterial',
                                    'Periodo']).agg({'Demanda':'sum',
                                                    }).reset_index()
        # agrupando datos
        # self.df.to_csv('Results/Agregado/Aexaminando_%s.csv'%self.par)
                                       
        self.df['CentroSuministrador'] = 'Osorno'
        

        self.df = self.df.merge(self.dfpredictors, 
                                on=['MesPeriodo','IdPeriodo',
                                    'Periodo','CentroSuministrador'],
                                how = 'left')
        
        self.df = self.df.dropna()
        
        self.df.MesPeriodo = self.df.MesPeriodo.map(str)
        self.dfpredictors.MesPeriodo = self.dfpredictors.MesPeriodo.map(str)
        
        df1 = self.df.MesPeriodo.str.get_dummies()
        df1.columns = ['is_' + col for col in df1.columns]
        self.df = pd.concat([self.df, df1], axis=1)
        self.temporal = df1.columns.tolist()
        
        df1 = self.dfpredictors.MesPeriodo.str.get_dummies()
        df1.columns = ['is_' + col for col in df1.columns]
        self.dfpredictors = pd.concat([self.dfpredictors, df1], axis=1)
        self.temporal = df1.columns.tolist()

        for key in self.dictpredictors.keys():
            self.dictpredictors[key] += self.temporal 
            
        for date_ in self.lastyear_name:
            if date_ not in self.dflastyear.columns.tolist():
                self.dflastyear[date_] = 0
                        
        dict_agg = {key : 'sum' for key in self.lastyear_name}
        pd_DDA_lastyear = self.dflastyear.groupby(['IdMaterial']).agg(dict_agg).reset_index()
        
        
        # agregando funcion de outliers
        
        self.df.insert(loc=5, column='outlier', value=0)
        self.df.insert(loc=0, column='position', value=self.df.index)

        general_outlier = self.general(self.df)
        indices = self.contador(self.df)
        if indices <= int(len(self.df)*0.08):
            self.df['Demanda'] = general_outlier['outlier']
        # print(general_outlier)
        
        # self.df.to_csv('Results/Agregado/read_%s.csv'%self.par)
        '''
        # agregando filtro de hampel
        
        ts = pd.Series(self.df['Demanda'])
        outlier_indices = hampel(ts, window_size=8, n=4)
        # print("Outlier Indices: ", outlier_indices)

        if len(outlier_indices) <= int(len(self.df)*0.08):
            ts_imputation = hampel(ts, window_size=8, n=4, imputation=True)
            self.df['Demanda'] = ts_imputation
        
        '''
        
        for sku in self.IdMaterial :
            
            try:
                                   
                m_ += 1
                
                df3 = self.df[(self.df.IdMaterial == sku) & 
                              (self.df['IdPeriodo'] < self.date_finish)].groupby(['IdPeriodo','IdMaterial']).agg({'Demanda':'sum'}).reset_index()
    
                print("==========SKU:%s, len:%s=========="%(sku,len(df3)))
                
                if len(df3) >  5 :
                    interval_time = len(df3['IdPeriodo'].unique().tolist())
                    interval_time += self.period 
                    date_start_sku = df3['IdPeriodo'].min()
                    
                    getdate = lambda x : date_start_sku + relativedelta(months=x)
        
                    predict_date_sku = [datetime.date(getdate(i).year, getdate(i).month, 1) for i in range(interval_time)]
                    
                
                    train_ = self.df[(self.df['IdMaterial'] == sku) & 
                                      (self.df['IdPeriodo'] <= self.date_start)]
                    
                    test_ = self.df[(self.df['IdMaterial'] == sku) & 
                                    (self.df['IdPeriodo'] > self.date_start) &
                                    (self.df['IdPeriodo'] < self.date_finish)]
                                        
                    dict_params = {}
                                               
                    print('\n',sku,'Porcentaje:%s'%(100*m_/len(self.IdMaterial)),'\n')
                    
                    with open('ReportScript/Forecast/Estado_ejecución_Pronóstico_%s.txt'%self.par, 'a') as f:
                        f.write('El sku actual es: %s, Porcentaje Completado: %s, Tiempo: %s(s) \n'%(sku,(100*m_/len(self.IdMaterial)) ,self.t0 - time.time()))   
                    f.close()
                    train = train_#[(train_['IdCeSum'] == suc)]
                                    
                    test  = self.dfpredictors[
                                              (self.dfpredictors['IdPeriodo'] > self.date_start) &
                                              (self.dfpredictors['IdPeriodo'] < self.date_finish)] #test_[(test_['IdCeSum']  == suc)]           
                                    
                    list_column_predict = ['IdPeriodo']
                    dict_predict = {}
                    RMSE = np.inf
                    MAPE = np.inf
                    dfpredict = pd.DataFrame()
                    error_models = {}
                    
                    dict_plot = {}
                    
                    idx_ = 0
                    vs_fore = {}
                    
                    pronostico = False
                    
                    for name_model in default_params.keys():
                        if name_model == 'rforest':
                            try:
                                params = self.ModelParamsRF.loc[sku,'Params']
                                params = params.replace('(','').replace(')','').replace("'",'')
                                params = [x.strip() for x in params.split(',')]
                                dict_params[sku,name_model] = 'Params'
                            except:
                                params = default_params[name_model] 
                                dict_params[sku,name_model] = 'Default'
        
                        elif name_model == 'lgbm':
                            try:
                                params = self.ModelParamsLGBM.loc[sku,'Params']
                                params = params.replace('(','').replace(')','').replace("'",'')
                                params = [x.strip() for x in params.split(',')]
                                dict_params[sku,name_model] = 'Params'
        
                            except:
                                params = default_params[name_model]
                                dict_params[sku,name_model] = 'Default'
        
                        elif name_model == 'svm':
                            try:
                                params = self.ModelParamsSVM.loc[sku,'Params']
                                params = params.replace('(','').replace(')','').replace("'",'')
                                params = [x.strip() for x in params.split(',')]
                                dict_params[sku,name_model] = 'Params'
        
                            except:
                                params = default_params[name_model]
                                dict_params[sku,name_model] = 'Default'
        
                        elif name_model == 'sarima':
                            try:
                                params = self.ModelParamsSarima.loc[sku,'Params']
                                params = params.replace('(','').replace(')','').replace("'",'')
                                params = [x.strip() for x in params.split(',')]
                                dict_params[sku,name_model] = 'Params'
        
                            except:
                                params = default_params[name_model]
                                dict_params[sku,name_model] = 'Default'    
                        
                        elif name_model == 'HoltWinteradd':
                            
                             try:
                                 params = self.ModelParamsHoltWinteradd.loc[sku,'Params']
                                 params = params.replace('(','').replace(')','').replace("'",'')
                                 params = [x.strip() for x in params.split(',')]
                                 dict_params[sku,name_model] = 'Params'
        
                             except:
                                 params = default_params[name_model]
                                 dict_params[sku,name_model] = 'Default'
                        
                        elif name_model == 'HoltWintermul':

                            try:
                                params = self.ModelParamsHoltWintermul.loc[sku,'Params']
                                params = params.replace('(','').replace(')','').replace("'",'')
                                params = [x.strip() for x in params.split(',')]
                                dict_params[sku,name_model] = 'Params'
        
                            except:
                                params = default_params[name_model]
                                dict_params[sku,name_model] = 'Default'
        
                        else:
                            params = default_params[name_model]
                            dict_params[sku,name_model] = 'Default'
                                    
                        if params[0] == '[Insuficiente Data]':
                            params = default_params[name_model]
                            dict_params[sku,name_model] = 'Default'
                        
                        try:
                            predict_, RMSE_, MAE, MAPE_, ddamean_, hist_predict = self.algorithm(name_model,params,train,test)
                            idx_ += 1
                            vs_fore[idx_]=[sku,name_model,RMSE_,MAE,MAPE_]
                            
                        except:
                            continue 
                    
                    sku = vs_fore[1][0]   
                    
                    df_vs_fore = pd.DataFrame.from_dict(vs_fore, orient='index')
                    df_vs_fore.to_csv('Results/Errores/errores_multi'+str(sku)+'.csv',sep = ';', encoding='ANSI', index=False, decimal=".")
                    
                    def obtener_max(diccionario):
                        MaxRMSE = 0
                        MaxMAE = 0
                        MaxMAPE = 0
                        for clave in diccionario:
                            for i in diccionario.values():
                                if diccionario[clave][2] > MaxRMSE:
                                    MaxRMSE = diccionario[clave][2]
                                elif diccionario[clave][3] > MaxMAE:
                                    MaxMAE = diccionario[clave][3]
                                elif diccionario[clave][4] > MaxMAPE:
                                    MaxMAPE = diccionario[clave][4]
                        return MaxRMSE,MaxMAE,MaxMAPE

                    def puntaje(diccionario,Divisor):
                        Puntaje = []
                        for clave in diccionario:
                            diccionario[clave][2] = diccionario[clave][2]/Divisor[0]
                            diccionario[clave][3] = diccionario[clave][3]/Divisor[1]
                            diccionario[clave][4] = diccionario[clave][4]/Divisor[2]
                            suma = diccionario[clave][2] + diccionario[clave][3] + diccionario[clave][4]
                            Puntaje.append(float(suma))
                        score = min(Puntaje)
                        winner = diccionario[int(Puntaje.index(score)+1)][1]
                        return winner
                     
                    
                    Divisor = obtener_max(vs_fore)
                    name_model =  puntaje(vs_fore,Divisor)
                    # print('name model aqui es: ',name_model)
                    if name_model == 'rforest':
                        try:
                            params = self.ModelParamsRF.loc[sku,'Params']
                            params = params.replace('(','').replace(')','').replace("'",'')
                            params = [x.strip() for x in params.split(',')]
                            dict_params[sku,name_model] = 'Params'
                        except:
                            params = default_params[name_model] 
                            dict_params[sku,name_model] = 'Default'
    
                    elif name_model == 'lgbm':
                        try:
                            params = self.ModelParamsLGBM.loc[sku,'Params']
                            params = params.replace('(','').replace(')','').replace("'",'')
                            params = [x.strip() for x in params.split(',')]
                            dict_params[sku,name_model] = 'Params'
    
                        except:
                            params = default_params[name_model]
                            dict_params[sku,name_model] = 'Default'
    
                    elif name_model == 'svm':
                        try:
                            params = self.ModelParamsSVM.loc[sku,'Params']
                            params = params.replace('(','').replace(')','').replace("'",'')
                            params = [x.strip() for x in params.split(',')]
                            dict_params[sku,name_model] = 'Params'
    
                        except:
                            params = default_params[name_model]
                            dict_params[sku,name_model] = 'Default'
    
                    elif name_model == 'sarima':
                        try:
                            params = self.ModelParamsSarima.loc[sku,'Params']
                            params = params.replace('(','').replace(')','').replace("'",'')
                            params = [x.strip() for x in params.split(',')]
                            dict_params[sku,name_model] = 'Params'
    
                        except:
                            params = default_params[name_model]
                            dict_params[sku,name_model] = 'Default'    
                    
                    elif name_model == 'HoltWinteradd':
                        
                         try:
                             params = self.ModelParamsHoltWinteradd.loc[sku,'Params']
                             params = params.replace('(','').replace(')','').replace("'",'')
                             params = [x.strip() for x in params.split(',')]
                             dict_params[sku,name_model] = 'Params'
    
                         except:
                             params = default_params[name_model]
                             dict_params[sku,name_model] = 'Default'
                    
                    elif name_model == 'HoltWintermul':

                        try:
                            params = self.ModelParamsHoltWintermul.loc[sku,'Params']
                            params = params.replace('(','').replace(')','').replace("'",'')
                            params = [x.strip() for x in params.split(',')]
                            dict_params[sku,name_model] = 'Params'
    
                        except:
                            params = default_params[name_model]
                            dict_params[sku,name_model] = 'Default'
    
                    else:
                        params = default_params[name_model]
                        dict_params[sku,name_model] = 'Default'
                                
                    if params[0] == '[Insuficiente Data]':
                        params = default_params[name_model]
                        dict_params[sku,name_model] = 'Default'
                    
                    predict_, RMSE_, MAE, MAPE_, ddamean_, hist_predict = self.algorithm(name_model,params,train,test)
                    error_models[sku,name_model] = RMSE_

                    
                    len_hist = len(hist_predict)
                    rand_ = np.random.randint([6],[len_hist-6], size = 2)
                    m1 = hist_predict[rand_[0]+6] - hist_predict[rand_[0]] 
                    m2 = hist_predict[rand_[1]+6] - hist_predict[rand_[1]]
                    
                    dict_ = {}
                    train2 = train['Demanda'].tolist()
                    
                    for i in range(len(train2)):
                        dict_[i] = [train2[i],hist_predict[i]]
                        
                    pddict = pd.DataFrame.from_dict(dict_,orient='index')
                                 
                    if (np.abs(m1 - m2) > 1e-3 and np.std(predict_) > .3):
                        
                        predict = predict_
                        RMSE = RMSE_
                        MAPE = MAPE_
                        ddamean = ddamean_
                        NameAlg = name_model
                        pronostico = True
                        
                    list_column_predict.append(name_model)
                    
                    for i in range(len(predict_)):
                        if i in dict_predict.keys():
                            dict_predict[i].append(predict_[i])
                        else:
                            dict_predict[i] = [predict_[i]]
                            
                    for i in range(len(hist_predict)):
                        if i in dict_plot.keys():
                            dict_plot[i].append(hist_predict[i])
                        else:
                            dict_plot[i] = [hist_predict[i]]

        
                    if pronostico == True:
                        np_hist_predict = np.array(hist_predict[:len(train)])
                        np_train = np.array(train)
                        
                        for i in range(len(hist_predict)):
                                                                            
                            predict_row = [predict_date_sku[i]] + dict_plot[i]
                            
                            s = pd.Series(predict_row)
            
                            dfpredict = dfpredict.append(s,ignore_index=True)
                        
                        dfpredict = dfpredict.set_axis(list_column_predict, axis=1)
                        
                        dfpredict['IdPeriodo'] = pd.to_datetime(predict_date_sku, format="%Y/%m")
                          
                        AcumProy  = sum(predict)
                        Trim1AcumProy  = sum(predict[:3])
                        Trim2AcumProy  = sum(predict[3:])
                        
                        ddamean = [np.mean(np.array(predict))]
            
                        umb = self.dfumb.loc[sku,'UMBase']
                        key = [sku, umb, len(train)]                    
                                    
                        Dda_lastyear = pd_DDA_lastyear.loc[(pd_DDA_lastyear['IdMaterial'] == sku)]
                        
                        Dda_lastyear = Dda_lastyear[self.lastyear_name].values.tolist()[0]
                        Dda_lastyear = [value for value in Dda_lastyear]
                                                
                        lastdate = min(self.date_lastyear)
                        t = 1
                        while len(Dda_lastyear) < 12 :
                            Dda_lastyear.insert(0,0)
                            t += 1
                            
                        if lastyearhead == False and t > 1:
                            for tt in range(1,t):
                                d = lastdate - relativedelta(months=tt)
                                self.lastyear_name.insert(0,'%s/%s'%(d.year,d.month))
                            lastyearhead = True
                                                        
                            
                        AcumEspj  = sum(Dda_lastyear[:6])
                        AcumPrev  = sum(Dda_lastyear[6:])
                        
                        trim_4 = sum(Dda_lastyear[:3])
                        trim_3 = sum(Dda_lastyear[3:6])
                        trim_2 = sum(Dda_lastyear[6:9])
                        trim_1 = sum(Dda_lastyear[9:])
                        
                        CotaInf   = (1-self.acceptance_tol_lb)*min(AcumEspj,AcumPrev)
                        CotaSup   = (1+self.acceptance_tol_ub)*max(AcumEspj,AcumPrev)
                            
                        def distance(value, lb, ub):
                            if ub >= value and value >= lb:
                                return 0
                            elif value > ub:
                                return value - ub
                            elif lb > value:
                                return ub - value
            
                        lastsemester_predict = np.concatenate((np.array(Dda_lastyear[6:]),
                                        np.array(predict)), axis=None)
            
                        DdaAnual  =   np.sum(lastsemester_predict)
                        MeanSupply =  np.mean(predict)
                        
                        Estado    = ('OK' if (CotaInf <= AcumProy) and
                                    (AcumProy <= CotaSup) 
                                    else ('COPIARESPEJO' if (CotaInf > AcumProy) else 'REVISAR'))
                        Presencia_ = [1 if Dda_lastyear[i] > 0 else 0 
                                      for i in range(len(Dda_lastyear))]
                        Presencia = sum(Presencia_)
                        
                        predict = [ round(predict[idx],2) for idx in range(len(predict)) ]
                        
                        current_row = (key + [NameAlg] + 
                                        predict + Dda_lastyear +
                                        ddamean + 
                                        [RMSE, 
                                        trim_4, trim_3, trim_2, trim_1,
                                        Trim1AcumProy, Trim2AcumProy,
                                        AcumProy, AcumEspj,
                                        AcumPrev, self.acceptance_tol_ub,
                                        self.acceptance_tol_lb, CotaInf, CotaSup,
                                        Estado, Presencia,
                                        DdaAnual,MeanSupply, MAPE
                                        ])
                                    
                        s = pd.Series(current_row)
                        dfsave = dfsave.append(s,ignore_index=True)
                        
                        df3 = df3[['IdPeriodo','Demanda']]
                        last_period = df3['IdPeriodo'].max()
                        first_period = df3['IdPeriodo'].min()
                        df3['IdPeriodo'] = pd.to_datetime(df3['IdPeriodo'], format="%Y/%m")
                 
                        df3 = df3.merge(dfpredict[['IdPeriodo',NameAlg]],
                                        on=['IdPeriodo'], how='right')
                        
                        df3 = df3.set_index('IdPeriodo')
                        
                        df3 = df3.rename(columns={'Demanda':'Real'})
                        
                                        
                        years = mdates.YearLocator()   # every year
                        months = mdates.MonthLocator()  # every month
                        years_fmt = mdates.DateFormatter('%Y/%m') #This is a format. Will be clear in Screenshot
                        
                        fig, axes = plt.subplots(figsize=(20,8))
                        
                        axes.vlines(self.date_start_predict, 0, df3['Real'].max(),
                                    linestyle='-.', color='m', 
                                    label='Start of forecast');
            
                        sns.lineplot(data=df3)
                        
                        axes.set(title = sku)
                        axes.xaxis.set_major_locator(months)
                        axes.xaxis.set_major_formatter(years_fmt)
                        axes.xaxis.set_minor_locator(months)
                        
                        up_span =  last_period + relativedelta(months=self.period)
                        lb_span =  last_period 
                        
                        while lb_span > first_period:
                            axes.axvspan(lb_span, up_span, facecolor = 'y', alpha = 0.25)
                            lb_span -= relativedelta(months=12)
                            up_span -= relativedelta(months=12)
                            
                        axes.axvspan(first_period, up_span, facecolor='y', alpha=0.25)
                            
                        plt.grid(axis = 'x', color = 'black', linestyle = '--', linewidth = 0.25)
                        plt.xticks(rotation=90, fontsize=7,)
                        plt.savefig("Plots/%s.jpg"%sku)
            except:
                with open('Error.txt', 'a') as f:
                    f.write('%s'%self.par) 
                f.close()
        
        # sku = vs_fore[1][0]   
        
        # df_vs_fore = pd.DataFrame.from_dict(vs_fore, orient='index')
        # df_vs_fore.to_csv('Results/Errores/errores_multi'+str(sku)+'.csv',sep = ';', encoding='ANSI', index=False, decimal=".")
        
        if len(dfsave) > 0:
        
            with open('ReportScript/Forecast/Estado_ejecución_Pronóstico_%s.txt'%self.par, 'a') as f:
                f.write('Ha finalizado el prónostico: %s(s), duración actual: %s(s) \n'%(time.time() -t0, time.time() - self.t0)) 
                f.write('Inicio de guardar data de pronóstico \n')
            f.close()
            
            list_column_name = (['IdMaterial','UMBase','Ndata', 'NameAlg'] +
                                self.period_proy +
                                self.lastyear_name +
                                ['DdaMean', 'RMSE'] +
                                self.trim_name  +
                                self.trim_proy +
                                ['AcumProy', 'AcumEsp', 'AcumPost', 'Factor_UB',
                                'Factor_LB',
                                'CotaInf', 'CotaSup', 'Estado', 'Presencia',
                                'DdaAnual','MeanSupply','MAPE'])
            
            dfsave = dfsave.set_axis(list_column_name, axis=1)
    
            dfsave['IdMaterial'] = dfsave['IdMaterial'].map(int)
    
            dfsave['CotaInf']   = (1-self.acceptance_tol_lb)*dfsave[['AcumEsp']].min(axis=1)
            dfsave['CotaSup']   = (1+self.acceptance_tol_ub)*dfsave[['AcumEsp']].max(axis=1)
    
            dfsave['Estado']    =  np.select(condlist = [(dfsave['CotaInf'] <=dfsave['AcumProy']) &
                                            (dfsave['AcumProy'] <= dfsave['CotaSup']),
                                            (dfsave['CotaInf'] > dfsave['AcumProy'])],
                                            choicelist=['OK', 'COPIARESPEJO'],
                                            default = 'REVISAR')
            
            def CopiarEspejo_init(estado, espejo, proyectado, idmaterial):
                            
                return espejo if estado == 'COPIARESPEJO' else proyectado 
            
            # dfsave[self.period_proy] = pd.DataFrame(dfsave.apply(lambda x: CopiarEspejo_init(x['Estado'],
            #                                                                                   x[self.lastyear_name[:6]],
            #                                                                                   x[self.period_proy],
            #                                                                                   x['IdMaterial']),
            #                                                                                   axis=1),
            #                                                                                   index=dfsave.index)
            
            dfsave['Editar']    =  np.select(condlist = [(dfsave['Estado'] == 'OK'),
                                            (dfsave['Estado'] == 'COPIARESPEJO')],
                                            choicelist=[dfsave['AcumProy'], dfsave['AcumEsp']],
                                            default = dfsave['AcumProy'])
    
    
            dfsave = dfsave.merge(self.dfabc[['IdMaterial','ABC']].drop_duplicates(),
                                  on= 'IdMaterial', how ='left')
            
            dfsave.to_csv('Results/Agregado/agregado_%s.csv'%self.par)
    
            def MeanStd_init(leadtime, supplies, mape): # Initial Inventory
                                
                period =  (leadtime+self.reviewperiod)/30

                i = 0
                mean, std = 0, 0
                
                while period > 0 and i<6:
                    prct = min(1,period)
                    mean += supplies[i]*prct
                    std  += ((1.25*min(mape,0.25)*supplies[i])**2)*prct
                    period -= prct
                    i += 1
                    
                std = np.sqrt(std)
                                                
                return mean, std
                
            columns_demand = ['AcumProy',
                              'AcumEsp',
                              'AcumPost',
                              'MeanSupply',
                              ] + self.trim_proy +  self.period_proy
            
            self.dflastyear = self.dflastyear.merge(dfsave[['IdMaterial',
                                                            'UMBase','MAPE']+columns_demand],
                                                    on = ['IdMaterial'],
                                                    how='left')
            for colname in columns_demand:
               self.dflastyear[colname] *= self.dflastyear['Porcentaje']
                    
            self.dflastyear['PeriodROP'] = self.dflastyear['LeadTime']/30
            self.dflastyear['PeriodROP'] = self.dflastyear['PeriodROP'].apply(np.ceil)
    
            self.dflastyear[['MeanSupply2','StdSupply2']] = self.dflastyear.apply(lambda x: MeanStd_init(x['LeadTime'],
                                                                       x[self.period_proy], x['MAPE']),
                                                                       axis=1).tolist()
            
            self.dflastyear['DdaAnual'] = self.dflastyear[self.period_proy +
                                                           self.lastyear_name[6:]].sum(axis=1)
            
            self.dflastyear['ABC'] = self.dflastyear['ABC'].fillna('B')
            self.dflastyear['ABC_Sucursal'] = self.dflastyear['ABC_Sucursal'].fillna('B')
                    
            self.dflastyear['DdaAnual'] = self.dflastyear[self.period_proy+self.lastyear_name[6:]].sum(axis=1)
            self.dflastyear['AcumEsp'] = self.dflastyear[self.lastyear_name[:6]].sum(axis=1)
            self.dflastyear['AcumPost'] = self.dflastyear[self.lastyear_name[6:]].sum(axis=1)
            
            print(self.trim_name[0],self.lastyear_name[:3])
            print(self.trim_name[1],self.lastyear_name[3:6])
            print(self.trim_name[2],self.lastyear_name[6:9])
            print(self.trim_name[3],self.lastyear_name[9:])
            
            self.dflastyear[self.trim_name[0]] = self.dflastyear[self.lastyear_name[:3]].sum(axis=1)
            self.dflastyear[self.trim_name[1]] = self.dflastyear[self.lastyear_name[3:6]].sum(axis=1)
            self.dflastyear[self.trim_name[2]] = self.dflastyear[self.lastyear_name[6:9]].sum(axis=1)
            self.dflastyear[self.trim_name[3]] = self.dflastyear[self.lastyear_name[9:]].sum(axis=1)
            
            # self.dflastyear['MeanSupply'] = self.dflastyear[self.period_proy].mean(axis=1)           
            
            self.dflastyear.to_csv('Results/PorSucursal/central_compra_%s.csv'%self.par)

            self.dflastyear['Provision'] = np.select(condlist=[self.dflastyear['IdCeSum'] =='D210'],
                                                          choicelist=['Centralizado'],
                                                          default= self.dflastyear['Provision'])

            central = self.dflastyear[self.dflastyear['Provision']=='Centralizado']
            
            #central['LeadTime'] = 7
            central['LeadTime'] = central['LeadTime'].fillna(7)
            
            central['ServiceLevel'] = np.select(condlist=[central['ABC_Sucursal']=='A',
                                                          central['ABC_Sucursal']=='B'],
                                                          choicelist=[self.levelconfidence_a,
                                                                      self.levelconfidence_b],
                                                          default= self.levelconfidence_c)

            central.to_csv('Results/Central_Abastecimiento/central_abastecimiento_%s.csv'%self.par)
            # Demanda agregada
            column_agg = (
                          self.period_proy + self.lastyear_name + self.trim_name + self.trim_proy)
            
            dict_agg = {key : 'sum' for key in column_agg}
                        
            
            central = self.dflastyear[self.dflastyear['Provision']=='Centralizado']            
        
            central   = central.groupby(['IdMaterial','UMBase']).agg(dict_agg).reset_index()
                       
            
            central['DdaAnual'] = central[self.period_proy+self.lastyear_name[6:]].sum(axis=1)
            central['AcumEsp'] = central[self.lastyear_name[:6]].sum(axis=1)
            central['AcumPost'] = central[self.lastyear_name[6:]].sum(axis=1)
            
            central[self.trim_name[0]] = central[self.lastyear_name[:3]].sum(axis=1)
            central[self.trim_name[1]] = central[self.lastyear_name[3:6]].sum(axis=1)
            central[self.trim_name[2]] = central[self.lastyear_name[6:9]].sum(axis=1)
            central[self.trim_name[3]] = central[self.lastyear_name[9:]].sum(axis=1)
            
            central['MeanSupply'] = central[self.period_proy].mean(axis=1)
            central = central.merge(dfsave[['IdMaterial','MAPE']],
                                    on='IdMaterial', how='left')
            
            central = central.merge(self.dfabc[['IdMaterial','ABC']].drop_duplicates(), 
                                    on='IdMaterial', how = 'left')
            
            central['ABC'] = central['ABC'].fillna('B')
            central['ServiceLevel'] = np.select(condlist=[central['ABC']=='A',
                                                          central['ABC']=='B'],
                                                          choicelist=[0.98, 0.95],
                                                          default=0.90)
                        
            idmaterials = list(central['IdMaterial'].unique())
            central = central.set_index(['IdMaterial'])
            
            for sku in idmaterials:
                try:
                    central.loc[sku,'LeadTime'] = dfM[(dfM['IdMaterial']==sku) & (dfM['IdCeSum']=='D210')]['LeadTime'].values[0] 
                except:
                    central.loc[sku,'LeadTime'] = 7
            central = central.reset_index() 
            
            if len(central) > 0:
                central[['MeanSupply2','StdSupply2']] = pd.DataFrame(central.apply(lambda x: MeanStd_init(x['LeadTime'],
                                                                      x[self.period_proy], x['MAPE']),
                                                                      axis=1).tolist(), index=central.index)
            
            
            central.to_csv('Results/Central_Compra/central_compra_%s.csv'%self.par)
            
            directo = self.dflastyear[self.dflastyear['Provision']=='Directo']
            
            directo['DdaAnual'] = directo[self.period_proy+self.lastyear_name[6:]].sum(axis=1)
            directo['AcumEsp'] = directo[self.lastyear_name[:6]].sum(axis=1)
            directo['AcumPost'] = directo[self.lastyear_name[6:]].sum(axis=1)

            directo[self.trim_name[0]] = directo[self.lastyear_name[:3]].sum(axis=1)
            directo[self.trim_name[1]] = directo[self.lastyear_name[3:6]].sum(axis=1)
            directo[self.trim_name[2]] = directo[self.lastyear_name[6:9]].sum(axis=1)
            directo[self.trim_name[3]] = directo[self.lastyear_name[9:]].sum(axis=1)

            directo['MeanSupply'] = directo[self.period_proy].mean(axis=1)
            directo = directo.merge(dfsave[['IdMaterial','MAPE']],
                                    on='IdMaterial',
                                    how='left')
            
            directo = directo[directo['IdCeSum'] != 'D210']
            directo['ABC_Sucursal'] = directo['ABC_Sucursal'].fillna('B')
            directo['ServiceLevel'] = np.select(condlist=[directo['ABC_Sucursal']=='A',
                                                          directo['ABC_Sucursal']=='B'],
                                                          choicelist=[0.98, 0.95],
                                                          default=0.90)
            
            directo.to_csv('Results/Directo/directo_%s.csv'%self.par)
            
            with open('ReportScript/Forecast/Estado_ejecución_Pronóstico_%s.txt'%self.par, 'a') as f:
                f.write('Ha finalizado de guardar la información: %s(s), duración actual: %s(s) \n'%(time.time() -t0, time.time() - self.t0)) 
            f.close()
        else:
            with open('Data_Insuficiente.txt', 'a') as f:
                f.write('%s-%s\n'%(self.par,sku)) 
            f.close()
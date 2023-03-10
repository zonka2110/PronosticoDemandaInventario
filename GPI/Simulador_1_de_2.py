import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import datetime 
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import calendar
import time

 

def generatePlotCost(ddaPronFull, df, op):

  StockSeg = df['StockSeg']
  ROP = df['ROP']

  indexDate = 0

  for i in range(len(ddaPronFull)):
    if(ddaPronFull.iloc[i]['Valor_Inv'] == min(ddaPronFull['Valor_Inv'])):
       indexDate = i
  
  x = ddaPronFull.iloc[indexDate]['ddaFecha']
  y = min(ddaPronFull['Valor_Inv'])

  plt.figure(figsize=(16,9))
  plt.text(x, y, 'Valor minimo : $'+str(int(y)), fontsize = 15)
  plt.plot(ddaPronFull['ddaFecha'], ddaPronFull['Valor_Inv'], label = 'Inv*Costo')
  plt.axhline(y = StockSeg*df['Costo'], xmin = 0, xmax = len(ddaPronFull), color = 'red', label = 'StockSeg*Costo')
  plt.axhline(y = ROP*df['Costo'], xmin = 0, xmax = len(ddaPronFull), color = 'green', label = 'ROP*Costo')
  plt.xticks(rotation = 90)
  plt.xticks(range(0, len(ddaPronFull['ddaFecha']), 3))
  plt.xlabel('ddaFecha')
  plt.ylabel('Valor Inventario')
  plt.tight_layout()
  plt.title(str(df['ID'])+'_'+df['Centro'], fontsize = 15)
  plt.legend(loc = 'best')


  namePlt = str(df['ID'])+'_'+df['Centro']

  if(op == 1):
    plt.savefig('ResultsCompraDirecta/PlotsValorInv/'+namePlt+'.jpg')

  elif(op == 2):
    plt.savefig('ResultsAbastecimientoCentral/PlotsValorInv/'+namePlt+'.jpg')

  elif(op == 3):
    plt.savefig('ResultsCompraCentral/PlotsValorInv/'+namePlt+'.jpg')
    
  else:
    print('error con op')
 


def generatePlotInv(ddaPronFull, df, op):

  StockSeg = df['StockSeg']
  ROP = df['ROP']

  indexPronInv = 0

  for i in range(len(ddaPronFull)):
    if(ddaPronFull.iloc[i]['Pron_Inv'] == min(ddaPronFull['Pron_Inv'])):
       indexPronInv = i
  
  x = ddaPronFull.iloc[indexPronInv]['ddaFecha']
  y = min(ddaPronFull['Pron_Inv'])

  plt.figure(figsize=(16,9))
  plt.text(x, y, 'Nivel minimo : '+str(round(y,3)), fontsize = 15)
  plt.plot(ddaPronFull['ddaFecha'], ddaPronFull['Pron_Inv'], label = 'Pron_Inv')
  plt.axhline(y = StockSeg, xmin = 0, xmax = len(ddaPronFull), color = 'red', label = 'StockSeg')
  plt.axhline(y = ROP, xmin = 0, xmax = len(ddaPronFull), color = 'green', label = 'ROP')
  plt.xticks(rotation = 90)
  plt.xticks(range(0, len(ddaPronFull['ddaFecha']), 3))
  plt.xlabel('ddaFecha')
  plt.ylabel('Pron Inventario')
  plt.tight_layout()
  plt.title(str(df['ID'])+'_'+df['Centro'], fontsize = 15)
  plt.legend(loc = 'best')
  #plt.show()

  namePlt = str(df['ID'])+'_'+df['Centro']

  if(op == 1):
    plt.savefig('ResultsCompraDirecta/PlotsInv/'+namePlt+'.jpg')

  elif(op == 2):
    plt.savefig('ResultsAbastecimientoCentral/PlotsInv/'+namePlt+'.jpg')

  elif(op == 3):
    plt.savefig('ResultsCompraCentral/PlotsInv/'+namePlt+'.jpg')
    
  else:
    print('error con op')



def getddaDiaria(month, df, df1):

  for j in range(len(df)):
    if(4 <= j <= 9):
      mes = df1.iloc[j].name[5:6]

      if(1 <= int(mes) <= 9):
        mes = '0'+mes 
          
      if(mes == month):
          
          if( (month == '01') or (month == '03') or (month == '05') or (month == '07') or (month == '08') or (month == '10') or (month == '12') ):
            ddaDiaria = float(df.iloc[j])/31
            
          elif( (month == '04') or (month == '06') or (month == '09') or (month == '11') ):
            ddaDiaria = float(df.iloc[j])/30

          elif( (month == '02') ): 
            year = int(df1.iloc[j].name[0:4])
          
            if( calendar.isleap(year) ):
              ddaDiaria = float(df.iloc[j])/29
            else:
              ddaDiaria = float(df.iloc[j])/28

          else:
            print('error al identificar el mes')
  
  return ddaDiaria



def editInv(ddaPronFull, pos, df):

  for i in range(len(ddaPronFull)):
    if(i == pos): 
       
      df1 = df.to_frame()
      month = ddaPronFull.iloc[i]['ddaFecha'][5:7]
      ddaPronFull.loc[i,'Pron_Inv'] = 0
      ddaPronFull.loc[i,'Valor_Inv'] = 0

      ddaDiaria = getddaDiaria(month, df, df1) 
      
      ddaPronFull.loc[i,'Pron_Inv'] = ddaPronFull.iloc[i]['Inv'] - ddaDiaria
      ddaPronFull.loc[i,'Valor_Inv'] = df['Costo']*ddaPronFull.iloc[i]['Inv']


  for i in range(len(ddaPronFull)):
    if(i == pos):
      temp = i+1
      for j in range(len(ddaPronFull)):
        if(temp <= j):
          ddaPronFull.loc[j,'Inv'] = 0
          ddaPronFull.loc[j,'Pron_Inv'] = 0
          ddaPronFull.loc[j,'Valor_Inv'] = 0
  
  
  for i in range(len(ddaPronFull)):
    if(i == pos):
      temp = i+1
      for j in range(len(ddaPronFull)):
        if(temp <= j):
          ddaPronFull.loc[j,'Inv'] = ddaPronFull.iloc[j-1]['Pron_Inv']
          month = ddaPronFull.iloc[j]['ddaFecha'][5:7]
          ddaDiaria = getddaDiaria(month, df, df1)
          ddaPronFull.loc[j,'Pron_Inv'] = ddaPronFull.loc[j,'Inv'] - ddaDiaria
          ddaPronFull.loc[j,'Valor_Inv'] = df['Costo']*ddaPronFull.iloc[j]['Inv']

  return ddaPronFull



def checkInv(ddaPronFull, df):

  next = 0
  index = 0
  flag = 0
  invRop = ddaPronFull.iloc[0]['Inv']
  pos = 0
  
  for i in range(len(df)):
    if(4 <= i <= 9):
      if( (df[i] == 0) and (df['Qopt'] == 0)):
        next = 1

  if(next != 1):
    for i in range(len(ddaPronFull)):
  
      if(i == 0):
        
        if(invRop < df['ROP']):
          pos = i + df['LT']

          if(pos <= (len(ddaPronFull)-1)):

            if(ddaPronFull.iloc[pos]['Inv'] < 0):
              ddaPronFull.loc[pos,'Inv'] = 0
              ddaPronFull.loc[pos,'Inv'] = df['Pedido']

            else:
              ddaPronFull.loc[pos,'Inv'] = ddaPronFull.iloc[pos]['Inv'] + df['Pedido']

            flag = 1
            index = i+1
            ddaPronFull = editInv(ddaPronFull, pos, df)

      else:
        
        if(i == index):
          
          if(flag == 1):
            ddaDiaria = getddaDiaria(ddaPronFull.iloc[i]['ddaFecha'][5:7], df, df.to_frame()) 
            invRop = ddaPronFull.iloc[i-1]['Inv'] - ddaDiaria + df['Pedido']
            flag = 0
          
          else:
            ddaDiaria = getddaDiaria(ddaPronFull.iloc[i]['ddaFecha'][5:7], df, df.to_frame()) 
            invRop = invRop - ddaDiaria + df['Qopt']


        else:
          ddaDiaria = getddaDiaria(ddaPronFull.iloc[i]['ddaFecha'][5:7], df, df.to_frame()) 
          if ddaPronFull.loc[i,'Inv']>=0:
            invRop = invRop - ddaDiaria

      
        if(invRop <= df['ROP']):
          pos = i + df['LT']

          if(pos <= (len(ddaPronFull)-1)):

            if(ddaPronFull.iloc[pos]['Inv'] < 0):
              ddaPronFull.loc[pos,'Inv'] = 0
              ddaPronFull.loc[pos,'Inv'] = df['Qopt']

            else:
              ddaPronFull.loc[pos,'Inv'] = ddaPronFull.iloc[pos]['Inv'] + df['Qopt']

            index = i+1
            ddaPronFull = editInv(ddaPronFull, pos, df)

  return ddaPronFull



def generateInventoryValue(ddaPronFull, df):

  ddaPronFull.insert(loc = 3, column = 'Valor_Inv', value = 0)

  for i in range(len(ddaPronFull)):
    ddaPronFull.loc[i,'Valor_Inv'] = df['Costo']*ddaPronFull.iloc[i]['Inv']
  
  ddaPronFull = checkInv(ddaPronFull, df)

  return ddaPronFull
  


def generateMonth(initialDdaPron, start_date, end_date, df):
  
  ddaMonth  = pd.DataFrame({'ddaFecha': [], 'Inv': [], 'Pron_Inv': []})
  
  date_Month = pd.date_range(start_date, end_date)

  for i in range(len(date_Month)):

    ddaMonth.loc[i,'ddaFecha'] = date_Month[i].strftime('%Y-%m-%d')
    ddaMonth.loc[i,'Inv'] = 0
    ddaMonth.loc[i,'Pron_Inv'] = 0

  ddaMonth.loc[0,'Inv'] = initialDdaPron['Inv']
  ddaMonth.loc[0,'Pron_Inv'] = initialDdaPron['Pron_Inv']

  for i in range(len(ddaMonth)):
    
    month = ddaMonth.iloc[i]['ddaFecha'][5:7]
    
    if( (month == '01') or (month == '03') or (month == '05') or (month == '07') or (month == '08') or (month == '10') or (month == '12') ):
      ddaDiaria = float(initialDdaPron['dda'])/31
    
    elif( (month == '04') or (month == '06') or (month == '09') or (month == '11') ):
      ddaDiaria = float(initialDdaPron['dda'])/30

    elif( (month == '02') ):
      year = int(str(ddaMonth.iloc[i]['ddaFecha'])[0:4])
      
      if( calendar.isleap(year) ):
        ddaDiaria = float(initialDdaPron['dda'])/29
      else:
        ddaDiaria = float(initialDdaPron['dda'])/28

    else:
      print('error al identificar el mes')
   

    if( (ddaMonth.iloc[i]['Inv'] == 0) and (ddaMonth.index[i] != 0) ):
    
      ddaMonth.loc[i,'Inv'] = ddaMonth.iloc[i-1]['Pron_Inv']
      ddaMonth.loc[i,'Pron_Inv'] = ddaMonth.loc[i,'Inv'] - ddaDiaria
      
  return ddaMonth 



def identifyMonth(ddaPron, df, op): 
  
  ddaPronFull  = pd.DataFrame({'ddaFecha': [], 'Inv': [], 'Pron_Inv': []})
  ddaMonth  = pd.DataFrame({'ddaFecha': [], 'Inv': [], 'Pron_Inv': []})

  for i in range(len(ddaPron)):

    month = str(ddaPron.iloc[i]['ddaFecha'])[5:7]

    if(month == '01'):

      start_date = str(ddaPron.iloc[i]['ddaFecha'])  
      end_date = start_date[0:5]+'01-31'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)
      
      if(start_date[8:10] == '31'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()
      
      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/28
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)
        
      
    elif(month == '02'):

      start_date = str(ddaPron.iloc[i]['ddaFecha'])  

      year = int(str(ddaPron.iloc[i]['ddaFecha'])[0:4])
      
      if( calendar.isleap(year) ):
        end_date = start_date[0:5]+'02-29'
      else:
        end_date = start_date[0:5]+'02-28'
        
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)
      
      if( calendar.isleap(year) ):
        if(start_date[8:10] == '29'):
          ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      elif(start_date[8:10] == '28'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']

      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()

      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/31
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)
    

    elif(month == '03'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'03-31'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '31'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()

      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/30 
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    elif(month == '04'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'04-30'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '30'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()

      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/31
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    elif(month == '05'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'05-31'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '31'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()
 
      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/30
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    elif(month == '06'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'06-30'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '30'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()

      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/31
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    elif(month == '07'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'07-31'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '31'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()

      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/31
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    elif(month == '08'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'08-31'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '31'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()
 
      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/30
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    elif(month == '09'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'09-30'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '30'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()

      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/31
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    elif(month == '10'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'10-31'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '31'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()

      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/30 
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    elif(month == '11'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'11-30'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '30'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()
 
      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/31
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    elif(month == '12'):
      
      start_date = str(ddaPron.iloc[i]['ddaFecha']) 
      end_date = start_date[0:5]+'12-31'
      ddaMonth = generateMonth(ddaPron.iloc[i], start_date, end_date, df)

      if(start_date[8:10] == '31'):
        ddaPron.loc[i+1,'Inv'] = ddaPron.iloc[i]['Pron_Inv']
      
      else:
        ddaPron.loc[i+1,'Inv'] = ddaMonth['Pron_Inv'].min()
 
      ddaPron.loc[i+1,'Pron_Inv'] = ddaPron.iloc[i+1]['Inv'] - float(ddaPron.iloc[i+1]['dda'])/31
      ddaPronFull = pd.concat([ddaPronFull, ddaMonth], ignore_index=True)


    else:

      print('error al identificar el mes') 


  ddaPronFull = generateInventoryValue(ddaPronFull, df)
  generatePlotCost(ddaPronFull, df, op) 
  generatePlotInv(ddaPronFull, df, op)

  return ddaPronFull



def firstPron_Inv(df, op):

  wb = openpyxl.Workbook()
  for i in range(len(df)):
    name_sheet = str(df.iloc[i]['ID'])+'_'+str(df.iloc[i]['Centro'])
    wb.create_sheet(index = i, title = name_sheet)
  
  if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
  
  for i in range(len(df)):
    ddaPron = pd.DataFrame({'ddaFecha': [], 'dda': []})
    
    for j in range(0,6): 
      txt = df.columns[j+4]
      format = '%Y/%m/%d'
      aux = datetime.datetime.strptime(txt,format)
      nuevaddaFecha = aux.strftime('%Y-%m-%d')
      ddaPron.loc[j, 'ddaFecha'] = nuevaddaFecha
      ddaPron.loc[j, 'dda'] = df.iloc[i][j+4]

    ddaPron.insert(loc = 2, column = 'Inv', value = 0)
    ddaPron.insert(loc = 3, column = 'Pron_Inv', value = 0)
    
    ddaPron.loc[0,'ddaFecha'] = df.iloc[i]['Fecha']
    month = str(ddaPron.iloc[0]['ddaFecha'])[5:7]

    if( (month == '01') or (month == '03') or (month == '05') or (month == '07') or (month == '08') or (month == '10') or (month == '12') ):
      ddaDiaria = float(ddaPron.loc[0,'dda'])/31
    
    elif( (month == '04') or (month == '06') or (month == '09') or (month == '11') ):
      ddaDiaria = float(ddaPron.loc[0,'dda'])/30

    elif( (month == '02') ):
      year = int(str(ddaPron.iloc[0]['ddaFecha'])[0:4])
      
      if( calendar.isleap(year) ):
        ddaDiaria = float(ddaPron.loc[0,'dda'])/29
      else:
        ddaDiaria = float(ddaPron.loc[0,'dda'])/28

    else:
      print('error al identificar el mes')

    flag = 0
    
    for k in range(len(ddaPron)):
      if( (ddaPron.iloc[k]['dda'] == 0) and df.iloc[i]['Qopt'] == 0):
        flag = 1
      
                        
    if(flag == 0):                                      
      ddaPron.loc[0,'Inv'] = df.iloc[i]['Inv']          
      ddaPron.loc[0,'Pron_Inv'] = ddaPron.loc[0,'Inv'] - ddaDiaria
    
    elif(flag == 1):
      ddaPron.loc[0,'Inv'] = 0
      ddaPron.loc[0,'Pron_Inv'] = ddaPron.loc[0,'Inv'] - ddaDiaria

    else:
      print('error flag')

    ddaPronFull = identifyMonth(ddaPron, df.iloc[i], op)
  
    nameSave = str(df.iloc[i]['ID'])+'_'+str(df.iloc[i]['Centro'])

    for x in wb.sheetnames:

      ws = wb[x]
      if(ws.title == nameSave):
        for r in dataframe_to_rows(ddaPronFull, index=True, header=True):
          ws.append(r)
        
        ws.delete_rows(2, 1)
    
    name = 'sku_'+str(df.iloc[0]['ID'])

    if(op == 1 or op == 2):
      if(op == 1):
        name1 = 'sku_CD_'+str(df.iloc[0]['ID'])
        wb.save('ResultsCDyAC/'+name1+'.xlsx')
      
      elif(op == 2):
        name2 = 'sku_AC_'+str(df.iloc[0]['ID'])
        wb.save('ResultsCDyAC/'+name2+'.xlsx')
      
      else:
        print('error con op')


    if(op == 1):
      wb.save('ResultsCompraDirecta/'+name+'.xlsx')
    
    elif(op == 2):
      wb.save('ResultsAbastecimientoCentral/'+name+'.xlsx')

    elif(op == 3):
      wb.save('ResultsCompraCentral/'+name+'.xlsx')
    
    else:
      print('error con op')



def list_sku(df):

  idList = []
  df_sku = df['ID']

  for i in range(len(df_sku)):

    if df_sku[i] not in idList:
      id = df_sku[i]
      idList.append(id)
  
  return idList



def findData(df, op):

  ids = list_sku(df)
  
  for i in range(len(ids)):
    
    ddaPronAux  = pd.DataFrame({'ID': [], 'Costo': [], 'Centro': [], 'UMB': [], df.columns[4]: [], df.columns[5]: [], df.columns[6]: [], df.columns[7]: [], 
                               df.columns[8]: [], df.columns[9]: [], 'Inv': [], 'ROP': [], 'StockSeg': [], 'LT': [], 'Qopt': [], 'Pedido': [], 'Fecha': []})
    
    for j in range(len(df)):

      if(df.iloc[j]['ID'] == ids[i]):
        ddaPronAux.loc[j,'ID'] = df.iloc[j]['ID']
        ddaPronAux.loc[j,'Costo'] = df.iloc[j]['Costo']
        ddaPronAux.loc[j,'Centro'] = df.iloc[j]['Centro']
        ddaPronAux.loc[j,'UMB'] = df.iloc[j]['UMB']
        ddaPronAux.loc[j,df.columns[4]] = df.iloc[j][df.columns[4]]
        ddaPronAux.loc[j,df.columns[5]] = df.iloc[j][df.columns[5]]
        ddaPronAux.loc[j,df.columns[6]] = df.iloc[j][df.columns[6]]
        ddaPronAux.loc[j,df.columns[7]] = df.iloc[j][df.columns[7]]
        ddaPronAux.loc[j,df.columns[8]] = df.iloc[j][df.columns[8]]
        ddaPronAux.loc[j,df.columns[9]] = df.iloc[j][df.columns[9]]
        ddaPronAux.loc[j,'Inv'] = df.iloc[j]['Inv']
        ddaPronAux.loc[j,'ROP'] = df.iloc[j]['ROP']
        ddaPronAux.loc[j,'StockSeg'] = df.iloc[j]['StockSeg']
        ddaPronAux.loc[j,'LT'] = df.iloc[j]['LT']
        ddaPronAux.loc[j,'Qopt'] = df.iloc[j]['Qopt']
        ddaPronAux.loc[j,'Pedido'] = df.iloc[j]['Pedido']
        ddaPronAux.loc[j,'Fecha'] = df.iloc[j]['Fecha'].date()

    ddaPronAux = ddaPronAux.astype({'ID':'int64', 'Costo':'int64', 'LT':'int64'})
    firstPron_Inv(ddaPronAux, op)

    

def fillData(df):

  for i in range(len(df)):
    
    if not (np.isnat(np.datetime64(str((df.iloc[i]['Fecha']))))):
      date = df.iloc[i]['Fecha']  
  
  for i in range(len(df)):

    if(np.isnat(np.datetime64(str((df.iloc[i]['Fecha']))))):
      df.loc[i,'Fecha'] = date

  for i in range(len(df)):

    list_Columns = ['ID', 'Costo', 'Centro', 'UMB', '2023/2/01', '2023/3/01', '2023/4/01', '2023/5/01', '2023/6/01', '2023/7/01', 'Inv', 'ROP', 'StockSeg', 'LT', 'Qopt', 'Pedido']

    for j in range(len(list_Columns)): 

      if(pd.isna(df.iloc[i][list_Columns[j]])):
        if( (list_Columns[j] == 'Centro') or (list_Columns[j] == 'UMB') ):
          df.loc[i,list_Columns[j]] = 'undefined'
        else:
          df.loc[i,list_Columns[j]] = 0

  return df

 

def main():
    
    inicio = time.time()

    # Compra_completa.xlsx es sku recibido y trabajado como entrada

    # Dataframe df1 'Compra_Directa'
    df1 = pd.read_excel('TestSku_2.xlsx', parse_dates = ['Fecha Stock'], sheet_name = 'Compra_Directa', usecols = [0,3,4,5,12,13,14,15,16,17,22,23,24,26,27,30,35])
    
    df1.rename(columns = {'IdMaterial':'ID', 'Costo\n Producto [$]':'Costo', df1.columns[4]:df1.columns[4][8:]+'/01', df1.columns[5]:df1.columns[5][8:]+'/01', 
                          df1.columns[6]:df1.columns[6][8:]+'/01', df1.columns[7]:df1.columns[7][8:]+'/01', df1.columns[8]:df1.columns[8][8:]+'/01', df1.columns[9]:df1.columns[9][8:]+'/01', 
                          'Stock Total':'Inv', 'Stock de Seguridad':'StockSeg', 'LeadTime\n [DIAS]':'LT', 'Q Ópt.':'Qopt', 'Solicitud de Pedido': 'Pedido', 'Fecha Stock':'Fecha'}, inplace = True)

    df1 = fillData(df1)
     


    # Dataframe df2 'Abastecimiento_Central'
    df2 = pd.read_excel('TestSku_2.xlsx', parse_dates = ['Fecha Stock'], sheet_name = 'Abastecimiento_Central', usecols = [0,3,4,5,12,13,14,15,16,17,22,23,24,26,27,30,35])

    df2.rename(columns = {'IdMaterial':'ID', 'Costo\n Producto [$]':'Costo', df2.columns[4]:df2.columns[4][8:]+'/01', df2.columns[5]:df2.columns[5][8:]+'/01', 
                          df2.columns[6]:df2.columns[6][8:]+'/01', df2.columns[7]:df2.columns[7][8:]+'/01', df2.columns[8]:df2.columns[8][8:]+'/01', df2.columns[9]:df2.columns[9][8:]+'/01', 
                          'Stock Total':'Inv', 'Stock de Seguridad':'StockSeg', 'LeadTime\n [DIAS]':'LT', 'Q Ópt.':'Qopt', 'Solicitud de Pedido': 'Pedido', 'Fecha Stock':'Fecha'}, inplace = True)

    df2 = fillData(df2)

    
    
    # Dataframe df3 'Compra_Central
    df3 = pd.read_excel('TestSku_2.xlsx', parse_dates = ['Fecha Stock'], sheet_name = 'Compra_Central', usecols = [0,3,4,5,12,13,14,15,16,17,22,23,24,26,27,30,35])
    
    df3.rename(columns = {'IdMaterial':'ID', 'Costo\n Producto [$]':'Costo', df3.columns[4]:df3.columns[4][8:]+'/01', df3.columns[5]:df3.columns[5][8:]+'/01', 
                          df3.columns[6]:df3.columns[6][8:]+'/01', df3.columns[7]:df3.columns[7][8:]+'/01', df3.columns[8]:df3.columns[8][8:]+'/01', df3.columns[9]:df3.columns[9][8:]+'/01', 
                          'Stock Total':'Inv', 'Stock de Seguridad':'StockSeg', 'LeadTime\n [DIAS]':'LT', 'Q Ópt.':'Qopt', 'Solicitud de Pedido': 'Pedido', 'Fecha Stock':'Fecha'}, inplace = True)
    
    df3 = fillData(df3)
    


    findData(df1, 1)
    findData(df2, 2)
    findData(df3, 3)

    fin = time.time()
    print('tiempo ejecucion: ',fin-inicio)
 
   
    
main()






 









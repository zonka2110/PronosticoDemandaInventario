"""
Inventory and distribution model.
"""

import numpy as np
import pandas as pd
from pyomo.util.infeasible import log_infeasible_constraints
from pyomo.opt import SolverStatus, TerminationCondition
from openpyxl import load_workbook
import logging
from dateutil.relativedelta import relativedelta

from scipy.stats import norm

import math as mt
import datetime
import os
import sys
import time

from ConnectionSQL import*

import psycopg2


import Data  # for path finding
p = str(Data.__path__)
l = p.find("'")
r = p.find("'", l+1)
data_path = p[l+1:r]

class Supplier:
    def __init__(self, 
                 setupcost = 40000,
                 discountrate = .22,
                 reviewperiod = 7,
                 name=None,
                 idSupplier=None,
                 idDataPredict=None):
        
        self.idSupplier = idSupplier
        self.idDataPredict = idDataPredict
        
        # Rutas de directorios
        
        self.path_datadirecto = 'Data/CentralDirecto.xlsx'


        
        # Rutas
        self.directo = 'Results/Directo.csv'.format(self.idDataPredict)
        self.porsucursal = 'Results/PorSucursal.csv'.format(self.idDataPredict)

        self.central_abastecimiento = 'Results/Central_Abastecimiento.csv'.format(self.idDataPredict)
        self.central_compra = 'Results/Central_Compra.csv'.format(self.idDataPredict)
        self.agregado = 'Results/Agregado.xlsx'.format(self.idDataPredict)
                
        # Paramétros de EOQ
        self.discountrate = discountrate # Tasa de descuento
        self.setupcost = setupcost # Costo de Setup
        self.reviewperiod = reviewperiod #Periodo de revisión
               
        # initialize columns of the report
        self.column_report1 = ['IdMaterial',
                                'NombreMaterial',
                                'ABC_Sucursal',
                                'ProductCost',
                                'Sucursal',
                                'UMBase']
        
        self.column_report2 = ['StockActual',
                               'OCenCurso', # Dividir interno y proveedor
                               'OCenInterno',
                               'StockTotal',
                               'ROP2',
                               'Stock de Seguridad',
                               'Deficit_Stock',
                               'LeadTime',
                               'Q',
                               'MesesComp',
                               'MesesInv',
                               'Abastecimiento2', #Solicitud de Pedido
                               # 'Opinion Sucursal',
                               # 'Analisis logistico',
                               # '% Llenado Stock Objetivo [Stock Total/ROP]',
                               # '% Llenado Stock Efectivo [Stock Actual/ROP]',	
                                # 'Meses Venta Stock Objetivo',
                                # 'Meses Venta Stock Actual',
                               # 'Pedido Suc (solped)',
                               # 'Valor Pedido',
                               # 'Peor Vencimiento stock actual',
                               'Stock en D210 (CDO)',
                               'Stock D190 (Osorno)',
                               'Stock Otras Sucursales',
                               # 'Minimo Compra Proveedor',
                               # 'Unidades por pallet',
                               # 'Ton/KG por pallet',                                                                                                                                                                                                                                                                                                                                                                                                                                                                      
                               'Fecha Stock',
                               'Fecha Transito']
            
    def getDataSupplier(self):
                                     
        self.dfabc  = pd.read_csv('Data/ABCxSucursal.csv', sep=",", decimal=',', index_col=0)
        
        suc = pd.read_csv('Data/Dictionary_Sucursal.csv', sep=",", decimal=',', index_col=0)
        suc = suc.rename(columns={'CentroSuministrador':'Sucursal'})
        suc['Sucursal'] = suc['Sucursal'].str.strip()
        
        self.FeatureItem = ReadSQL('coo_consumos')
        self.FeatureItem['IdMaterial'] = self.FeatureItem['IdMaterial'].map(int)

        self.FeatureItem = self.FeatureItem.dropna()
            
        self.FeatureItem['IdPeriodo']      = self.FeatureItem['Periodo'].map(str) + '-' + self.FeatureItem['MesPeriodo'].map(str)
        self.FeatureItem["IdPeriodo"]      = pd.to_datetime(self.FeatureItem['IdPeriodo'], format="%Y/%m")
        
        self.FeatureItem = self.FeatureItem[['IdMaterial', 'IdPeriodo', 'IdCeSum', 'MntCosto', 'CtdadUMBase']]
                     
        # Obetenemos la fecha del último año
        date_current  = self.FeatureItem['IdPeriodo'].max()
        date_lastyear = date_current - relativedelta(months=12)

        # Se filtra la data del último año, y se agrupa por SKU y Sucursal
        self.FeatureItem = self.FeatureItem[(self.FeatureItem['IdPeriodo'] <= date_current) & 
                (self.FeatureItem['IdPeriodo'] >= date_lastyear)]
        
        self.FeatureItem['MntCosto']      = pd.to_numeric(self.FeatureItem['MntCosto'])
        self.FeatureItem['CtdadUMBase']   = pd.to_numeric(self.FeatureItem['CtdadUMBase'])
        
        self.FeatureItem = self.FeatureItem.groupby(['IdMaterial']).agg({'MntCosto':'sum',
                                                     'CtdadUMBase':'sum'}
                                                    ).reset_index()
            
        self.FeatureItem['CostoUnitario'] = self.FeatureItem['MntCosto']/self.FeatureItem['CtdadUMBase']
                        
        self.FeatureItem = self.FeatureItem.groupby(['IdMaterial']).agg({
                                                             'CostoUnitario' :'mean'}).reset_index()
                                                                                                                                     
        self.FeatureItem.rename(columns={
                                       'CostoUnitario':'ProductCost'},
                                        inplace = True)

        # #%% SetupCost 45 dolares (810 pesos chilenos)
        self.FeatureItem['SetupCost'] = self.setupcost#45*810
        
        # #%% InventoryCost: Costo de Compra promedio * 22%
        self.FeatureItem['ProductCost'] = self.FeatureItem['ProductCost'].round(0)

        df2 = self.FeatureItem[self.FeatureItem['ProductCost'] == 0]
        df2.to_csv('dfceros.csv')
        
        self.FeatureItem = self.FeatureItem[self.FeatureItem['ProductCost'] != 0]

        
        self.FeatureItem['InventoryCost'] = self.FeatureItem['ProductCost']*self.discountrate
                           
        return 0#self.dfsupply

    def CalcularROP(self,skufilter):
        
        def ROP_init(service_level,dda_mean,dda_std,leadtime): # Initial Inventory
            _z = norm.ppf(service_level)
            
            try:
                t_mean = dda_mean
            except:
                t_mean = 0
            try:
                t_std = dda_std*_z  
            except:
                t_std = 0
                
            T = t_mean + t_std

            return T
        
        def Compra_init(rop, StockTotal, q):
            return q + rop - StockTotal if rop >= StockTotal else 0

        def Q_init(dda_mean,cost_setup,cost_inv):
            return np.sqrt((2*dda_mean*cost_setup)/cost_inv)  
        
        def AggData_init(df):
            df['LeadTime']     = df['LeadTime'].map(float)
            df['ServiceLevel'] = df['ServiceLevel'].map(float)
            df['MeanSupply2']  = df['MeanSupply2'].map(float)
            df['StdSupply2']   = df['StdSupply2'].map(float)
            df['MeanSupply']   = df['MeanSupply'].map(float)
        
            df['Cambio']       = df['Cambio'].map(float)
            df['StockActual']  = df['StockActual'].map(float)
            df['OCenCurso']  = df['OCenCurso'].map(float)
            df['OCenInterno']  = df['OCenInterno'].map(float)
            
            df['DdaAnual']     = df['DdaAnual'].map(float)

            df['OCenCurso']   = df['OCenCurso'].fillna(0)
            df['OCenInterno']   = df['OCenInterno'].fillna(0)
            df['StockActual'] = df['StockActual'].fillna(0)

            df['OCenCurso']   = df['OCenCurso']
            df['OCenInterno']   = df['OCenInterno']
            df['StockActual'] = df['StockActual']

            df['StockTotal'] = df['OCenCurso'] + df['StockActual'] + df['OCenInterno']
            
            df['Cambio'] = pd.to_numeric(df['Cambio'] , downcast='float')
            df['MeanSupply'] = pd.to_numeric(df['MeanSupply'] , downcast='float')
                        
            df['DdaAnual'] += (df['Cambio']-1.0)*6.0*df['MeanSupply']
    
            df['MeanSupply2'] = df['MeanSupply2']*df['Cambio']
            df['StdSupply2']  = df['StdSupply2']*df['Cambio']            
                                
            df['ROP2'] = df.apply(lambda x: ROP_init(x.ServiceLevel, x.MeanSupply2,
                                                               x.StdSupply2, x.LeadTime),
                                                                   axis=1)
            df['Stock de Seguridad'] =  df.StdSupply2*norm.ppf(df.ServiceLevel)
            
            df['Q'] = df.apply(lambda x: Q_init(x.DdaAnual, x.SetupCost,
                                                          x.InventoryCost), axis=1)

            df['MesesComp']  = 12*df['Q']/df['DdaAnual']
            df['MesesComp'] = df['MesesComp'].round(1)
            
            df['MesesInv']  = 12*df['StockTotal']/df['DdaAnual']
            df['MesesInv'] = df['MesesInv'].round(1)
            
            df['Abastecimiento2'] = df.apply(lambda x: Compra_init(x.ROP2, x.StockTotal,
                                                x.Q) , axis = 1) 
            
            df['Deficit_Stock'] =  df.apply(lambda x: max(x.ROP2- x.StockTotal,0),
                                                      axis = 1)
            
            df['% Llenado Stock Objetivo [Stock Total/ROP]'] = df.apply(lambda x: x.StockTotal/x.ROP2 if x.ROP2>0
                                                                else np.inf,
                                                       axis = 1)
            
            df['% Llenado Stock Efectivo [Stock Actual/ROP]'] = df.apply(lambda x: x.StockActual/x.ROP2 if x.ROP2>0
                                                                else np.inf,
                                                       axis = 1)
            
            df['Meses Venta Stock Objetivo'] =  12*df['ROP2']/df['DdaAnual']
            df['Meses Venta Stock Actual']  =  12*df['StockTotal']/df['DdaAnual']
                                    
            # Llena cooprinsem
            df['Opinion Sucursal'] = ''
            df['Analisis logistico'] = ''
            df['Pedido Suc (solped)'] = ''
            df['Valor Pedido'] = ''
            df['Peor Vencimiento stock actual'] = ''
            df['Minimo Compra Proveedor'] = ''
            df['Unidades por pallet'] = ''
            df['Ton/KG por pallet'] = ''
            
            df = df.merge(self.stock_suc[['IdMaterial',
                                                    'D210',
                                                    'D190',
                                                    'Otras']], how='left', on=['IdMaterial'])
            
            df = df.rename(columns = {'D210':  'Stock en D210 (CDO)',
                                                'D190':  'Stock D190 (Osorno)',
                                                'Otras': 'Stock Otras Sucursales',
                                                })
            
            return df
        
        #%% Obtener data para el cualculo de abastecimiento y orden de compra                
        
        # CentralDirecto = ReadSQL('coo_centraldirecto')
        # CentralDirecto = CentralDirecto[['IdMaterial','Sucursal','Provision']]
        
        pred = pd.read_csv(self.porsucursal,  sep=',', decimal='.', index_col = 0)

        
        self.SaveSheet2(pred[pred.LeadTime.isnull()][['IdMaterial','LeadTime']], 'Sin LeadTime')
       
        StockActual = ReadSQL('coo_stock')
        
        descr_ = StockActual[['IdMaterial',
                              'NombreMaterial']].drop_duplicates()
        
        StockActual = StockActual[['IdMaterial',
                                   'Sucursal',
                                   'CtdadStockUMB',
                                   'fechaproceso']].rename(columns={'CtdadStockUMB':'StockActual',
                                                                    'fechaproceso':'Fecha Stock'})
        
        StockActual = StockActual.groupby(['IdMaterial',
                                           'Sucursal',
                                           'Fecha Stock']).agg({'StockActual':'sum'}).reset_index()
        
        self.stock_suc       = pd.pivot_table(StockActual,
                                              values  = 'StockActual',
                                              columns = 'Sucursal',
                                              index   = ['IdMaterial'],
                                              fill_value = 0)
        
        self.stock_suc["Otras"] = self.stock_suc.sum(axis=1) - self.stock_suc['D210'] - self.stock_suc['D190']
        self.stock_suc = self.stock_suc.reset_index()
        
        OCenCurso = ReadSQL('coo_entransito')
        OCenCurso['CtdadPendienteUMB'] = OCenCurso.apply(lambda x: max(0,x.CtdadPendienteUMB), axis=1)
        OCenCurso= OCenCurso[['IdMaterial',
                              'IdCentro',
                              'CtdadPendienteUMB',
                              'fechaproceso']].rename(columns={'CtdadPendienteUMB':'OCenCurso',
                                                               'IdCentro':'Sucursal',
                                                               'fechaproceso':'Fecha Transito'})
        
        OCenCurso.loc[OCenCurso["OCenCurso"] < 0, "OCenCurso"] = 0
        OCenCurso = OCenCurso.groupby(['IdMaterial','Sucursal','Fecha Transito']).agg({'OCenCurso':'sum'}).reset_index()

        OCenCursoInterno = ReadSQL('coo_zmb5tq')
        OCenCursoInterno['CtdadInternoUMB'] = OCenCursoInterno.apply(lambda x: max(0,x.CtdadInternoUMB), axis=1)
        OCenCursoInterno= OCenCursoInterno[['IdMaterial',
                              'IdCentro',
                              'CtdadInternoUMB',
                              'fechaproceso']].rename(columns={'CtdadInternoUMB':'OCenInterno',
                                                               'IdCentro':'Sucursal',
                                                               'fechaproceso':'Fecha Interno'})
        
        # OCenCursoInterno.loc[OCenCursoInterno["OCenCurso"] < 0, "OCenCurso"] = 0
        OCenCursoInterno = OCenCursoInterno.groupby(['IdMaterial','Sucursal','Fecha Interno']).agg({'OCenInterno':'sum'}).reset_index()
        OCenCursoInterno['IdMaterial'] = OCenCursoInterno['IdMaterial'].map(int)
                        
        revisar = pd.read_excel(self.agregado)#,  sep=',', decimal='.', index_col = 0)
        
        dfresult = revisar.merge(self.FeatureItem, on='IdMaterial', how='left')
        dfresult = dfresult[dfresult.ProductCost.isnull()]
        
        self.SaveSheet2(dfresult[['IdMaterial','ProductCost']], 'Sin Costo')
        
        revisar = revisar[['IdMaterial','Editar','AcumProy']]
        revisar['Editar'] = revisar['Editar'].map(float)
        revisar['AcumProy'] = revisar['AcumProy'].map(float)
        revisar['Cambio'] = revisar['Editar']/revisar['AcumProy']
        revisar = revisar[['IdMaterial','Cambio']]       
                
        #%% Calculo de Orden de Compra Directo
        dfresult = pd.read_csv(self.directo,  sep=',', decimal='.', index_col = 0)
        if skufilter != []:
            
            dfresult= dfresult[dfresult['IdMaterial'].isin(skufilter)]
            
        dfresult = dfresult.rename(columns={'Sucursal':'Ciudad',
                                          'IdCeSum':'Sucursal'})
                        
        if len(dfresult) != 0:
            
            dfresult = pd.merge(dfresult, OCenCurso, on=['IdMaterial','Sucursal'], how='left')
            dfresult = pd.merge(dfresult, OCenCursoInterno, on=['IdMaterial','Sucursal'], how='left')
            dfresult.to_excel('dfresult.xlsx')
            OCenCursoInterno.to_excel('OCenCursoInterno.xlsx')
            dfresult = pd.merge(dfresult, StockActual, on=['IdMaterial','Sucursal'], how='left')
            dfresult = dfresult.merge(self.FeatureItem, on='IdMaterial', how='left')
            dfresult = dfresult.merge(revisar, on='IdMaterial', how='left')
            dfresult = dfresult.merge(descr_, on='IdMaterial', how='left')
            
            dfresult = dfresult[(dfresult['ProductCost'] >= 0) & 
                                (dfresult['LeadTime'] >= 0)]
            
            
                        
            dfresult = AggData_init(dfresult)
            
            columns_ = dfresult.columns.tolist()

            ddaproy = list(filter(lambda x: 'DdaProy' in x, columns_))
            ddatrimreal = list(filter(lambda x: 'Trim 2' in x, columns_))
           
            self.column_report1 = list(map(lambda x: x.replace('ABC_Sucursal', 'ABC'), self.column_report1))
            
            ddatrimproy = list(filter(lambda x: 'Trim Proy' in x, columns_))

            for trim in ddatrimproy:
               dfresult[trim] = dfresult['Cambio']*dfresult[trim]
               
            for dda in ddaproy:
               dfresult[dda] = dfresult['Cambio']*dfresult[dda]
           
            dfresult = dfresult[self.column_report1 +
                            ddatrimreal + ddatrimproy +ddaproy + ['DdaAnual'] +
                            self.column_report2 
                            ]
            
            dfresult = dfresult.drop_duplicates()
            
            dfresult = self.AdjustDataFrame(dfresult)
            
            self.SaveSheet2(dfresult, 'Compra_Directa')
              
        #%% Generación de Abastecimiento caso Centralizado.
        dfresult = pd.read_csv(self.central_abastecimiento,  sep=',', decimal='.', index_col = 0)
        if skufilter != []:
            dfresult= dfresult[dfresult['IdMaterial'].isin(skufilter)]
        dfresult = dfresult.rename(columns={'Sucursal':'Ciudad',
                                        'IdCeSum':'Sucursal'})
        if len(dfresult) != 0:
            
            
            dfresult = pd.merge(dfresult, OCenCurso, on=['IdMaterial','Sucursal'], how='left')
            dfresult = pd.merge(dfresult, OCenCursoInterno, on=['IdMaterial','Sucursal'], how='left')
            dfresult = pd.merge(dfresult, StockActual, on=['IdMaterial','Sucursal'], how='left')

            dfresult = dfresult.merge(self.FeatureItem, on='IdMaterial', how='left')
            dfresult = dfresult.merge(revisar, on='IdMaterial', how='left')
            dfresult = dfresult.merge(descr_, on='IdMaterial', how='left')

            dfresult = dfresult[(dfresult['ProductCost'] >= 0) & 
                                (dfresult['LeadTime'] >= 0)]
            
            dfresult = AggData_init(dfresult)

            columns_ = dfresult.columns.tolist()
            
            
            ddaproy = list(filter(lambda x: 'DdaProy' in x, columns_))
            ddatrimreal = list(filter(lambda x: 'Trim 2' in x, columns_))
           
            self.column_report1 = list(map(lambda x: x.replace('ABC_Sucursal', 'ABC'), self.column_report1))
            
            ddatrimproy = list(filter(lambda x: 'Trim Proy' in x, columns_))

            for trim in ddatrimproy:
               dfresult[trim] = dfresult['Cambio']*dfresult[trim]
           
            dfresult = dfresult[self.column_report1 +
                            ddatrimreal + ddatrimproy +ddaproy + ['DdaAnual'] +
                            self.column_report2 
                            ]
            
            dfresult = dfresult.drop_duplicates()
            
            dfresult  = self.AdjustDataFrame(dfresult)
            
            self.SaveSheet2(dfresult, 'Abastecimiento_Central')

         #%% Calculo de Orden de  dfresultizado
        dfresult = pd.read_csv(self.central_compra,  sep=',', decimal='.', index_col = 0)
        if skufilter != []:
            dfresult= dfresult[dfresult['IdMaterial'].isin(skufilter)]
 
        if len(dfresult) != 0:
                     
             dfresult = dfresult.merge(self.FeatureItem, on='IdMaterial', how='left')
             dfresult = dfresult.merge(revisar, on='IdMaterial', how='left')
             dfresult = dfresult.merge(descr_, on='IdMaterial', how='left')

             dfresult = dfresult[(dfresult['ProductCost'] >= 0) & 
                                (dfresult['LeadTime'] >= 0)]
                         
             dfresult = dfresult.merge(OCenCurso[OCenCurso['Sucursal'] == 'D210'][['IdMaterial',
                                                                                 'OCenCurso',
                                                                                 'Fecha Transito']],
                                     on = 'IdMaterial', how = 'left')

             dfresult = dfresult.merge(OCenCursoInterno[OCenCursoInterno['Sucursal'] == 'D210'][['IdMaterial',
                                                                                 'OCenInterno',
                                                                                 'Fecha Interno']],
                                     on = 'IdMaterial', how = 'left')

             
             dfresult = dfresult.merge(StockActual[StockActual['Sucursal'] == 'D210'][['IdMaterial',
                                                                                     'StockActual',
                                                                                     'Fecha Stock']],
                                     on = 'IdMaterial', how = 'left')
             
             dfresult = AggData_init(dfresult)
             
             dfresult['Sucursal'] = 'D210'
             
             columns_ = dfresult.columns.tolist()
                          
             ddaproy = list(filter(lambda x: 'DdaProy' in x, columns_))
             ddatrimreal = list(filter(lambda x: 'Trim 2' in x, columns_))
           
             self.column_report1 = list(map(lambda x: x.replace('ABC_Sucursal', 'ABC'), self.column_report1))
            
             ddatrimproy = list(filter(lambda x: 'Trim Proy' in x, columns_))

             for trim in ddatrimproy:
                dfresult[trim] = dfresult['Cambio']*dfresult[trim]
           
             dfresult = dfresult[self.column_report1 +
                            ddatrimreal + ddatrimproy +ddaproy + ['DdaAnual'] +
                            self.column_report2 
                            ]
             
             dfresult = dfresult.drop_duplicates()
             
             dfresult = self.AdjustDataFrame(dfresult) 
             
             self.SaveSheet2(dfresult, 'Compra_Central')
        
    #%% Creating Excel Output ## ---------------------------------->

    def AdjustDataFrame(self, df):
        
        def RoundbyUMB(umb, list_): # Initial Inventory
            if umb == 'UN  ':
                list_ = [round(value,0) for value in list_]
            else:
                list_ = [round(value,2) for value in list_]
            return list_
                
        df['StockActual']   = df['StockActual'].round(2)
        df['OCenCurso']     = df['OCenCurso'].round(2)
        df['OCenInterno']     = df['OCenInterno'].round(2)
        df['ROP2']          = df['ROP2'].round(2)
        df['Deficit_Stock'] = df['Deficit_Stock'].round(2)
        
        df['Q']             = df['Q'].round(2)
        df['Deficit_Stock'] = df['Deficit_Stock'].round(2)
        df['Abastecimiento2'] = df['Abastecimiento2'].round(2)
        
        columns_ = df.columns.tolist()
        ddatrim = list(filter(lambda x: 'Trim' in x, columns_))
        
        for trim in ddatrim:
            df[trim] = df[trim].round(2)
        
        ddaproy = list(filter(lambda x: 'DdaProy' in x, columns_))
        
        for dda in ddaproy:
            df[dda] = df[dda].round(2)
        
        df = df.rename(columns={'IdMaterial':'IdMaterial Cooprinsem',
                                'Sucursal': 'Centro',
                                'UMBase': 'UMB',
                                'StockActual' :'Stock en Sucursal',
                                'OCenCurso' : 'Stock Transito OC ', # Dividir interno y proveedor
                                'OCenInterno' : 'Stock Transito Interno', # Dividir interno y proveedor
                                'StockTotal' : 'Stock Total', # Dividir interno y proveedor
                                'ROP2':'ROP',
                                'Deficit_Stock':'Deficit de Stock (*)',
                                'Q': 'Q Ópt.',
                                'Abastecimiento2': 'Solicitud de Pedido', 
                                'Opinion Sucursal': 'Opinión sucursal',
                                'Analisis logistico': 'Análisis logístico',
                                'ProductCost'    :'Costo\n Producto [$]',
                                'LeadTime'       :'LeadTime\n [DIAS]',
                                'MesesComp'       :'Meses\n Stock de Compra [12*Q/Dda Anual]',
                                'MesesInv'       :'Meses\n Stock de Inventario [12*StockTotal/DdaAnual]',
                                'NombreMaterial' : 'Nombre Material',
                                'DdaAnual' : 'Dda. Anual [his+proy]'
                                })
       
        return df
          
    def SaveSheet2(self, dfsave, namesheet):
                                 
        path = r"Results/CompraAbastecimiento{}.xlsx".format(self.idSupplier)  
        try:
            book = load_workbook(path)
            writer = pd.ExcelWriter(path, engine = 'openpyxl')
            writer.book = book
            sheets = book.sheetnames
            if namesheet in sheets:
                std=book[namesheet]
                book.remove(std)
                dfsave.to_excel(writer, sheet_name=namesheet, index=False)
                writer.save()
            else:
                dfsave.to_excel(writer, sheet_name=namesheet, index=False)
                writer.save()
        except:
            with pd.ExcelWriter(path) as writer:
                dfsave.to_excel(writer, sheet_name=namesheet, index=False)   
        
    def run(self,skufilter):
        self.getDataSupplier()
        self.CalcularROP(skufilter)
        

if len(sys.argv)>1:
    _idSupplier = sys.argv[1]
else:
    _idSupplier = '04e06004-prueba'

if len(sys.argv)>2:
    _idDataPredict = sys.argv[2]
else:
    _idDataPredict = '04e06004-1f30-11ed-9718-42010a8e0002'


if len(sys.argv)>3:
    _setupcost = int(sys.argv[3])
    _discountrate = int(sys.argv[4])/100
    _reviewperiod = int(sys.argv[5])
    _family = sys.argv[6]
else:
    _setupcost = 40000
    _discountrate = .22
    _reviewperiod = 7
    _family = 'todos'

sp = Supplier(setupcost=_setupcost,discountrate=_discountrate,reviewperiod=_reviewperiod,idSupplier=_idSupplier,idDataPredict=_idDataPredict,name=0)#sys.argv[1])
t0 = time.time()

if _family != 'todos':
  _family = _family.split(',')[:-1]
  print(_family)
  df = ReadSQL('coo_sku_x_familia')
  dat = df.loc[df['familia'].isin(_family)]
  x=dat['id_material'].tolist()
else:
  x = []

print(x)
sp.run(x)  
print('Tiempo de Duración %s'%(time.time()-t0))
